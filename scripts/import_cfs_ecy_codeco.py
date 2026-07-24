#!/usr/bin/env python3
"""Idempotent importer for the CFS-ECY CODECO gate-movement feeds (module 13).

Loads the two JNPA CODECO workbooks into the ADDITIVE table from migration 0027:
  * CFS-CODECO.xlsx -> core.cfs_ecy_movement (facility_type='CFS')
  * ECY-CODECO.xlsx -> core.cfs_ecy_movement (facility_type='ECY')

Each file has three columns: "Container Number", "Timestamp", "Mode".
Purely additive — it NEVER touches cargo / empty_container / vehicle / driver /
transporter / auth tables. It only appends raw gate events; the container_number
soft-links to core.cargo BY VALUE (no FK).

Per row:
  * container_number : trimmed, validated with jnpa_shared.iso6346 (iso_valid flag)
  * timestamp        : parsed DD/MM/YYYY HH:MM, stamped IST (Asia/Kolkata) -> timestamptz
  * mode             : In -> IN, Out -> OUT
  * facility_type    : derived from the source filename (CFS / ECY)
A row is INVALID (not imported) only if it lacks a container number, a parseable
timestamp, or a recognised mode. Exact-duplicate gate events are dropped by the
unique constraint via ON CONFLICT DO NOTHING (idempotent re-runs).

Usage:
    python scripts/import_cfs_ecy_codeco.py --dry-run          # parse+validate, no DB
    POSTGRES_DSN='postgresql+asyncpg://postgres:TempPass123!@localhost:5433/postgres' \
        .venv/bin/python scripts/import_cfs_ecy_codeco.py      # live upsert (+ ensures schema)
Options: --data-dir PATH, --dsn, --dry-run, --limit N, --no-ensure.
"""
from __future__ import annotations

import argparse
import asyncio
import datetime as dt
import os
import sys
from collections import Counter
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))
sys.path.insert(0, str(_ROOT / "shared"))

from jnpa_shared.iso6346 import is_valid_container_no  # noqa: E402

DEFAULT_DATA_DIR = "/Users/pandurangdhage/Downloads/Digital Twin/Data/13-CFS-ECY"
DEFAULT_DSN = os.environ.get(
    "POSTGRES_DSN",
    "postgresql+asyncpg://postgres:TempPass123!@localhost:5433/postgres",
)
BATCH = 500

# JNPA operates in IST. The CODECO timestamps carry no timezone; we stamp them as
# Asia/Kolkata (UTC+5:30) so the timestamptz column stores the correct instant.
IST = dt.timezone(dt.timedelta(hours=5, minutes=30))

# (source filename, facility_type)
FILES = (("CFS-CODECO.xlsx", "CFS"), ("ECY-CODECO.xlsx", "ECY"))
_COLS = ("Container Number", "Timestamp", "Mode")


# --- normalization -----------------------------------------------------------
def clean_text(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    v = str(raw).strip()
    return v or None


def norm_mode(raw: Any) -> Optional[str]:
    """In -> IN, Out -> OUT (case-insensitive). Anything else -> None (invalid)."""
    v = (clean_text(raw) or "").upper()
    return v if v in ("IN", "OUT") else None


def parse_ts(raw: Any) -> Optional[dt.datetime]:
    """Parse the CODECO timestamp (DD/MM/YYYY HH:MM) into an IST-aware datetime.
    Accepts an already-parsed Excel datetime too."""
    if raw is None:
        return None
    if isinstance(raw, dt.datetime):
        return raw if raw.tzinfo else raw.replace(tzinfo=IST)
    s = str(raw).strip()
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%d-%m-%Y %H:%M"):
        try:
            return dt.datetime.strptime(s, fmt).replace(tzinfo=IST)
        except ValueError:
            continue
    return None


def clean_row(raw: Dict[str, Any], facility: str, source_file: str
              ) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    """Return (record | None, reason_if_invalid)."""
    cn = clean_text(raw.get("Container Number"))
    if not cn:
        return None, "missing_container"
    cn = cn.upper()
    ts = parse_ts(raw.get("Timestamp"))
    if ts is None:
        return None, "bad_timestamp"
    mode = norm_mode(raw.get("Mode"))
    if mode is None:
        return None, "bad_mode"
    return {
        "facility_type": facility,
        "container_number": cn,
        "iso_valid": bool(is_valid_container_no(cn)),
        "event_ts": ts,
        "mode": mode,
        "source": "CODECO",
        "source_file": source_file,
    }, None


# --- workbook ----------------------------------------------------------------
def load_sheet(xlsx: str, cols, limit: Optional[int]) -> List[Dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    idx = {h: i for i, h in enumerate(header)}
    missing = [c for c in cols if c not in idx]
    if missing:
        wb.close()
        raise SystemExit(f"FATAL: {xlsx} missing columns {missing}")
    out: List[Dict[str, Any]] = []
    for n, values in enumerate(it):
        if limit is not None and n >= limit:
            break
        if not any(v not in (None, "") for v in values):
            continue  # skip fully-blank rows
        out.append({c: (values[idx[c]] if idx[c] < len(values) else None) for c in cols})
    wb.close()
    return out


_INSERT = """
INSERT INTO core.cfs_ecy_movement
    (facility_type, container_number, iso_valid, event_ts, mode, source, source_file)
VALUES
    (:facility_type, :container_number, :iso_valid, :event_ts, :mode, :source, :source_file)
ON CONFLICT ON CONSTRAINT uq_cfs_ecy_movement DO NOTHING
"""


async def import_records(records: List[Dict[str, Any]], dsn: str) -> int:
    """Chunked idempotent insert. Returns the number of NEW rows actually inserted
    (ON CONFLICT DO NOTHING makes duplicates a no-op)."""
    from sqlalchemy import text

    from jnpa_shared.db import get_engine

    engine = get_engine(dsn)
    stmt = text(_INSERT)
    inserted = 0
    for i in range(0, len(records), BATCH):
        async with engine.begin() as conn:
            for rec in records[i:i + BATCH]:
                res = await conn.execute(stmt, rec)
                inserted += (res.rowcount or 0)
    return inserted


def build_report(data_dir: str, limit: Optional[int]) -> Dict[str, Any]:
    per_file: Dict[str, Any] = {}
    valid: List[Dict[str, Any]] = []
    invalid: List[Dict[str, Any]] = []
    iso_invalid = 0
    reasons: Counter = Counter()
    for fname, facility in FILES:
        path = Path(data_dir) / fname
        if not path.exists():
            raise SystemExit(f"FATAL: data file not found: {path}")
        raw_rows = load_sheet(str(path), _COLS, limit)
        seen: set = set()
        f_valid = f_invalid = f_dupe = 0
        for raw in raw_rows:
            rec, reason = clean_row(raw, facility, fname)
            if rec is None:
                f_invalid += 1
                reasons[reason] += 1
                invalid.append({"file": fname, "row": raw, "reason": reason})
                continue
            key = (rec["facility_type"], rec["container_number"],
                   rec["event_ts"], rec["mode"])
            if key in seen:
                f_dupe += 1  # in-file exact duplicate (also blocked by the DB constraint)
                continue
            seen.add(key)
            if not rec["iso_valid"]:
                iso_invalid += 1
            f_valid += 1
            valid.append(rec)
        per_file[facility] = {"file": fname, "raw_rows": len(raw_rows),
                              "valid": f_valid, "invalid": f_invalid,
                              "in_file_duplicates": f_dupe}
    return {"per_file": per_file, "valid": valid, "invalid": invalid,
            "iso_invalid": iso_invalid, "reasons": dict(reasons)}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--data-dir", default=DEFAULT_DATA_DIR)
    ap.add_argument("--dsn", default=DEFAULT_DSN)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--no-ensure", action="store_true",
                    help="skip the boot-ensure DDL (assume migration 0027 already applied)")
    args = ap.parse_args()

    rep = build_report(args.data_dir, args.limit)

    inserted = None
    if not args.dry_run:
        async def run() -> int:
            if not args.no_ensure:
                from gateway.cfs_ecy_ext import ensure_cfs_ecy_schema
                await ensure_cfs_ecy_schema(args.dsn)
            return await import_records(rep["valid"], args.dsn)
        inserted = asyncio.run(run())

    total_valid = len(rep["valid"])
    total_invalid = len(rep["invalid"])
    dupes = sum(f["in_file_duplicates"] for f in rep["per_file"].values())

    print("\n" + "=" * 66)
    print("CFS-ECY CODECO IMPORT" + ("  [DRY-RUN — no DB writes]" if args.dry_run else ""))
    print("=" * 66)
    for facility, f in rep["per_file"].items():
        print(f"  {facility} ({f['file']}): raw={f['raw_rows']}  valid={f['valid']}  "
              f"invalid={f['invalid']}  in_file_dupes={f['in_file_duplicates']}")
    print("-" * 66)
    print(f"  importable (valid, deduped)   : {total_valid}")
    print(f"  invalid (not imported)        : {total_invalid}  reasons={rep['reasons']}")
    print(f"  in-file duplicate rows skipped: {dupes}")
    print(f"  ISO-6346 invalid (flagged)    : {rep['iso_invalid']}")
    if args.dry_run:
        print("  rows inserted                 : n/a (dry-run)")
    else:
        print(f"  rows inserted (new)           : {inserted}")
        print(f"  duplicates skipped by DB      : {total_valid - (inserted or 0)}")
    if rep["invalid"]:
        print("\n  INVALID SAMPLES:")
        for r in rep["invalid"][:10]:
            print(f"    {r}")
    print("=" * 66 + "\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
