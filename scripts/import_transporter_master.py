#!/usr/bin/env python3
"""Idempotent importer for the official Transport Master dataset.

Loads ``TransporterDetails.xlsx`` (JNPA source system) into the EXISTING
``jnpa.transporters`` entity. This is additive: it only touches the Transport
Master columns added by migration 0025 and NEVER touches the blacklist,
vehicle-mapping or validation tables. Existing operator-entered rows (which have
no ``source_company_id``) are left alone.

Idempotency key: the source ``company_id`` -> ``jnpa.transporters.source_company_id``
(a UNIQUE index). Re-running upserts by that key, so:
  * a brand-new company_id      -> INSERTED
  * a known company_id, changed -> UPDATED
  * a known company_id, same    -> SKIPPED (no write)

Cleaning / normalization applied per row (see ``clean_row``):
  * mobile  - strip non-digits, drop 91/0 country/trunk prefixes, keep 10-digit
  * email   - basic validation; a value that isn't an email is quarantined (NULL)
  * designation - trim + canonical Title-Case (OWNER/owner -> Owner, PROPRITOR ->
                  Proprietor, ...)
  * address - trimmed; empty -> NULL
  * company names - duplicates are reported (informational); company_id stays the
                    key, so duplicate names never collide.

A record is INVALID (not imported) only when it lacks a usable key or name:
missing/non-numeric company_id, or empty company_name.

Usage:
    # Dry-run: parse + clean + classify, no DB writes (needs only openpyxl)
    python scripts/import_transporter_master.py --dry-run

    # Live upsert against the running stack (Postgres on localhost:5433)
    POSTGRES_DSN='postgresql+asyncpg://postgres:jnpa_pw@localhost:5433/postgres' \
        .venv/bin/python scripts/import_transporter_master.py

Options: --xlsx PATH, --dsn DSN, --dry-run, --limit N, --report PATH (JSON).
Exit code 0 on success (even with invalid rows); 2 on a fatal error.
"""
from __future__ import annotations

import argparse
import asyncio
import json
import os
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))
sys.path.insert(0, str(_ROOT / "shared"))

DEFAULT_XLSX = (
    "/Users/pandurangdhage/Downloads/Digital Twin/Data/"
    "11-Transport Data/TransporterDetails.xlsx"
)
DEFAULT_DSN = os.environ.get(
    "POSTGRES_DSN",
    "postgresql+asyncpg://postgres:jnpa_pw@localhost:5433/postgres",
)

# Source-column -> our field name. Header row of the sheet must contain these.
_COLUMNS = {
    "company_id": "source_company_id",
    "user_user_id": "source_user_id",
    "company_name": "name",
    "contactPersonName": "contact_person",
    "designation": "designation",
    "email": "email",
    "mobile_number": "mobile",
    "address": "address",
    "company_document1": "doc_type",
    "company_document_file1": "doc_file",
}

# Canonical spelling fixes applied AFTER Title-casing the raw designation.
_DESIGNATION_FIX = {
    "Propritor": "Proprietor",
    "Propriter": "Proprietor",
    "Proprietior": "Proprietor",
    "Managar": "Manager",
    "Manger": "Manager",
    "Maneger": "Manager",
    "Ceo": "CEO",
    "Md": "MD",
    "Hr": "HR",
}

_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


# --- normalization helpers ---------------------------------------------------
def clean_mobile(raw: Any) -> Tuple[Optional[str], bool]:
    """Return (normalized 10-digit mobile or None, was_fixed_or_dropped)."""
    if raw is None:
        return None, False
    digits = re.sub(r"\D", "", str(raw))
    changed = digits != str(raw).strip()
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    elif len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]
    if len(digits) == 10:
        return digits, changed
    return None, True  # unusable length -> quarantine


def clean_email(raw: Any) -> Tuple[Optional[str], bool]:
    """Return (lowercased valid email or None, was_quarantined)."""
    if raw is None:
        return None, False
    val = str(raw).strip()
    if _EMAIL_RE.match(val):
        return val.lower(), False
    return None, True  # not an email (e.g. address shifted into the field)


def clean_designation(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    val = str(raw).strip()
    if not val:
        return None
    titled = val.title()
    return _DESIGNATION_FIX.get(titled, titled)


def clean_text(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    val = str(raw).strip()
    return val or None


def clean_int(raw: Any) -> Optional[int]:
    if raw is None:
        return None
    try:
        return int(str(raw).strip())
    except (ValueError, TypeError):
        return None


def clean_row(raw: Dict[str, Any]) -> Tuple[Optional[Dict[str, Any]], List[str]]:
    """Clean one source row. Returns (record | None, issues).

    ``record is None`` => INVALID (not importable). ``issues`` lists per-field
    fixes/quarantines for the validation report even when the record is kept.
    """
    issues: List[str] = []
    source_company_id = clean_int(raw.get("company_id"))
    name = clean_text(raw.get("company_name"))
    if source_company_id is None:
        issues.append("missing_or_bad_company_id")
        return None, issues
    if not name:
        issues.append("missing_company_name")
        return None, issues

    mobile, mob_changed = clean_mobile(raw.get("mobile_number"))
    if mobile is None and raw.get("mobile_number") not in (None, ""):
        issues.append("mobile_quarantined")
    elif mob_changed:
        issues.append("mobile_normalized")

    email, email_quarantined = clean_email(raw.get("email"))
    if email_quarantined and raw.get("email") not in (None, ""):
        issues.append("email_quarantined")

    designation = clean_designation(raw.get("designation"))
    if designation is None and raw.get("designation") not in (None, ""):
        issues.append("designation_empty")

    address = clean_text(raw.get("address"))
    if address is None:
        issues.append("address_missing")

    record = {
        "source_company_id": source_company_id,
        "source_user_id": clean_int(raw.get("user_user_id")),
        "name": name,
        "contact_person": clean_text(raw.get("contactPersonName")),
        "designation": designation,
        "email": email,
        "mobile": mobile,
        "address": address,
        "doc_type": clean_text(raw.get("company_document1")),
        "doc_file": clean_text(raw.get("company_document_file1")),
    }
    return record, issues


# --- workbook loading --------------------------------------------------------
def load_rows(xlsx_path: str, limit: Optional[int]) -> List[Dict[str, Any]]:
    import openpyxl  # local import so --help works without the dep

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    idx = {h: i for i, h in enumerate(header)}
    missing = [c for c in _COLUMNS if c not in idx]
    if missing:
        wb.close()
        raise SystemExit(f"FATAL: source columns not found in sheet: {missing}")
    rows: List[Dict[str, Any]] = []
    for n, values in enumerate(it):
        if limit is not None and n >= limit:
            break
        rows.append({c: (values[idx[c]] if idx[c] < len(values) else None)
                     for c in _COLUMNS})
    wb.close()
    return rows


# --- upsert ------------------------------------------------------------------
_UPSERT = """
INSERT INTO jnpa.transporters AS t
    (source_company_id, source_user_id, name, contact_person, designation,
     email, mobile, address, doc_type, doc_file, contact, status)
VALUES
    (:source_company_id, :source_user_id, :name, :contact_person, :designation,
     :email, :mobile, :address, :doc_type, :doc_file,
     CAST(:contact AS jsonb), 'ACTIVE')
ON CONFLICT (source_company_id) DO UPDATE SET
    source_user_id = EXCLUDED.source_user_id,
    name           = EXCLUDED.name,
    contact_person = EXCLUDED.contact_person,
    designation    = EXCLUDED.designation,
    email          = EXCLUDED.email,
    mobile         = EXCLUDED.mobile,
    address        = EXCLUDED.address,
    doc_type       = EXCLUDED.doc_type,
    doc_file       = EXCLUDED.doc_file,
    contact        = EXCLUDED.contact,
    updated_at     = now()
WHERE (t.source_user_id, t.name, t.contact_person, t.designation, t.email,
       t.mobile, t.address, t.doc_type, t.doc_file, t.contact)
  IS DISTINCT FROM
      (EXCLUDED.source_user_id, EXCLUDED.name, EXCLUDED.contact_person,
       EXCLUDED.designation, EXCLUDED.email, EXCLUDED.mobile, EXCLUDED.address,
       EXCLUDED.doc_type, EXCLUDED.doc_file, EXCLUDED.contact)
RETURNING (xmax = 0) AS inserted
"""


def _contact_json(rec: Dict[str, Any]) -> str:
    """Populate the legacy `contact` jsonb so existing read paths stay consistent."""
    contact = {k: rec[k] for k in ("email", "mobile", "address")
               if rec.get(k) is not None}
    if rec.get("mobile"):
        contact["phone"] = rec["mobile"]
    return json.dumps(contact)


async def upsert_all(records: List[Dict[str, Any]], dsn: str) -> Dict[str, int]:
    from jnpa_shared.db import execute_returning, ping

    if not await ping(dsn=dsn):
        raise SystemExit("FATAL: database not reachable at the given DSN")
    tally = {"inserted": 0, "updated": 0, "skipped": 0}
    for rec in records:
        params = dict(rec)
        params["contact"] = _contact_json(rec)
        row = await execute_returning(_UPSERT, params, dsn=dsn)
        if row is None:
            tally["skipped"] += 1
        elif row.get("inserted"):
            tally["inserted"] += 1
        else:
            tally["updated"] += 1
    return tally


# --- report ------------------------------------------------------------------
def build_report(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    valid: List[Dict[str, Any]] = []
    invalid: List[Dict[str, Any]] = []
    issue_counts: Counter = Counter()
    name_counts: Counter = Counter()
    for raw in rows:
        rec, issues = clean_row(raw)
        for i in issues:
            issue_counts[i] += 1
        if rec is None:
            invalid.append({"company_id": raw.get("company_id"),
                            "company_name": raw.get("company_name"),
                            "reasons": issues})
        else:
            valid.append(rec)
            name_counts[rec["name"].lower()] += 1
    dup_names = {n: c for n, c in name_counts.items() if c > 1}
    return {
        "total_rows": len(rows),
        "valid": valid,
        "invalid": invalid,
        "issue_counts": dict(issue_counts),
        "duplicate_name_groups": len(dup_names),
        "duplicate_name_rows": sum(dup_names.values()),
    }


def print_summary(report: Dict[str, Any], tally: Optional[Dict[str, int]],
                  dry_run: bool) -> None:
    print("\n" + "=" * 68)
    print("TRANSPORT MASTER IMPORT" + ("  [DRY-RUN — no DB writes]" if dry_run else ""))
    print("=" * 68)
    print(f"  source rows read      : {report['total_rows']}")
    print(f"  importable (valid)    : {len(report['valid'])}")
    print(f"  invalid (not imported): {len(report['invalid'])}")
    print("\n  IMPORT SUMMARY")
    if dry_run:
        print(f"    inserted (projected): {len(report['valid'])}")
        print("    updated             : n/a (dry-run)")
        print("    skipped             : n/a (dry-run)")
    else:
        assert tally is not None
        print(f"    inserted            : {tally['inserted']}")
        print(f"    updated             : {tally['updated']}")
        print(f"    skipped (no change) : {tally['skipped']}")
    print(f"    invalid             : {len(report['invalid'])}")
    print("\n  VALIDATION / CLEANING")
    for k in sorted(report["issue_counts"]):
        print(f"    {k:22}: {report['issue_counts'][k]}")
    print(f"    duplicate_name_groups : {report['duplicate_name_groups']}"
          f" ({report['duplicate_name_rows']} rows)")
    if report["invalid"]:
        print("\n  INVALID SAMPLES (max 10):")
        for r in report["invalid"][:10]:
            print(f"    company_id={r['company_id']!r} name={r['company_name']!r}"
                  f" reasons={r['reasons']}")
    print("=" * 68 + "\n")


def main() -> int:
    ap = argparse.ArgumentParser(description="Import Transport Master dataset")
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--dsn", default=DEFAULT_DSN)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--report", default=None, help="write full report JSON here")
    args = ap.parse_args()

    if not Path(args.xlsx).exists():
        print(f"FATAL: xlsx not found: {args.xlsx}", file=sys.stderr)
        return 2

    rows = load_rows(args.xlsx, args.limit)
    report = build_report(rows)

    tally: Optional[Dict[str, int]] = None
    if not args.dry_run:
        tally = asyncio.run(upsert_all(report["valid"], args.dsn))

    print_summary(report, tally, args.dry_run)

    if args.report:
        out = {k: v for k, v in report.items() if k != "valid"}
        out["valid_count"] = len(report["valid"])
        out["db_tally"] = tally
        Path(args.report).write_text(json.dumps(out, indent=2, default=str))
        print(f"report written -> {args.report}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
