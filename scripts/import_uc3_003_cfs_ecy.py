#!/usr/bin/env python3
"""UC3-003 — load the REAL CFS/ECY CODECO gate logs into core.container_event.

Source (the customer's corpus, Data/13-CFS-ECY):
    ECY-CODECO.xlsx   961 rows   Empty Container Yard gate log
    CFS-CODECO.xlsx   968 rows   Container Freight Station gate log

Both workbooks carry three columns — "Container Number", "Timestamp" (DD/MM/YYYY
HH:MM, IST, no timezone) and "Mode" (In / Out) — and the facility is carried by
the FILENAME, not by a column. Each source row becomes exactly one
core.container_event:

    file  mode   ->  event_type   location_type   direction
    ECY   Out        ECY_OUT      ECY             O
    ECY   In         ECY_IN       ECY             I
    CFS   In         CFS_IN       CFS             I
    CFS   Out        CFS_OUT      CFS             O

Nothing else is modelled: core.container_event already has every column this
feed needs, so no parallel event table is created (see migration 0133, which
adds only indexes and two derived views).

WHAT THIS IMPORTER WILL NOT DO
------------------------------
The ECY feed is deliberately unpaired — 529 OUT events (01–12 Jul) and 432 IN
events (12–26 Jul) in two date-disjoint blocks that share not one container.
The CFS feed carries one exact-duplicate row. Those are the source's own
anomalies and they are **detected, never patched**: every source row is imported
verbatim (duplicates included, multiplicity preserved), no timestamp or
container number is altered, no event is invented to complete a pair, and every
finding is written to the existing core.dq_issue ledger as a grouped row.

IDEMPOTENCY
-----------
There is no unique constraint on core.container_event — and there must not be,
because the corpus legitimately contains the same gate event twice. So the
importer is *multiplicity-aware*: for each distinct
(container_no, event_ts, event_type) it inserts only
``source_count - already_in_db_count`` rows. First run inserts 1929; every later
run inserts 0 and rewrites nothing.

Usage:
    python scripts/import_uc3_003_cfs_ecy.py --dry-run           # parse + analyse only
    POSTGRES_DSN='postgresql+asyncpg://…' \
        .venv/bin/python scripts/import_uc3_003_cfs_ecy.py       # live import
Options: --corpus PATH, --dsn, --dry-run, --json.
"""
from __future__ import annotations

import argparse
import asyncio
import datetime as dt
import json
import os
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

_ROOT = Path(__file__).resolve().parents[1]
sys.path[:0] = [str(_ROOT), str(_ROOT / "shared")]

from jnpa_shared.iso6346 import is_valid_container_no  # noqa: E402

# JNPA operates in IST; the CODECO timestamps carry no timezone.
IST = dt.timezone(dt.timedelta(hours=5, minutes=30))

# The three columns both workbooks must have, in the customer's own spelling.
COLUMNS: Tuple[str, str, str] = ("Container Number", "Timestamp", "Mode")

# Timestamp formats accepted, most-likely first. The corpus uses the first one;
# the others exist so a re-drop with a different Excel locale still parses
# rather than silently dropping rows.
TS_FORMATS = ("%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%d-%m-%Y %H:%M",
              "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M")


@dataclass(frozen=True)
class Feed:
    """One source workbook and how its rows map into core.container_event."""

    filename: str
    facility: str                     # location_type: ECY | CFS
    # source_table marks the provenance of every row this feed writes. These are
    # the customer's OWN markers (their seed already labels these events this
    # way), so an environment that was seeded from their dump is recognised as
    # already-imported and the run is a no-op instead of a duplication.
    source_table: str
    ingest_path: str                  # core.ingest_file.path (natural key)
    expected_rows: int                # the count the brief states; reported, never enforced


FEEDS: Tuple[Feed, ...] = (
    Feed("ECY-CODECO.xlsx", "ECY", "staging ecy_codeco",
         "Data/13-CFS-ECY/ECY-CODECO.xlsx", 961),
    Feed("CFS-CODECO.xlsx", "CFS", "staging cfs_codeco",
         "Data/13-CFS-ECY/CFS-CODECO.xlsx", 968),
)

INGEST_SOURCE_SYSTEM = "CFS-ECY"      # the customer's own source_system tag
DQ_SOURCE_TABLE = "core.container_event"

# Where to look for the corpus when --corpus is not given. The first directory
# that holds BOTH workbooks wins.
CORPUS_CANDIDATES = (
    os.environ.get("JNPA_CORPUS_DIR"),
    "data/13-CFS-ECY",
    "../Data/13-CFS-ECY",
    "~/Downloads/Data/13-CFS-ECY",
    "~/Downloads/Digital Twin/Data/13-CFS-ECY",
    "~/Downloads/Digital Twin Data Corpus - Updated/Data/13-CFS-ECY",
)

BATCH = 500


class SourceError(RuntimeError):
    """The corpus is missing or does not have the shape this importer needs."""


# ============================== source parsing ==============================
@dataclass(frozen=True)
class SourceEvent:
    """One source row, normalised but never altered."""

    container_no: str
    event_ts: dt.datetime
    event_type: str
    location_type: str
    direction: str
    source_table: str
    iso_valid: bool
    raw_container: str
    raw_timestamp: str
    raw_mode: str
    source_file: str
    source_row: int                   # 1-based row number in the sheet, header excluded

    @property
    def key(self) -> Tuple[str, dt.datetime, str]:
        """Identity of the *gate event* — what idempotency is measured against."""
        return (self.container_no, self.event_ts, self.event_type)


def find_corpus(explicit: Optional[str]) -> Path:
    """Locate the directory holding both CODECO workbooks."""
    candidates: List[str] = [explicit] if explicit else list(CORPUS_CANDIDATES)
    tried: List[str] = []
    for cand in candidates:
        if not cand:
            continue
        base = Path(cand).expanduser()
        if not base.is_absolute():
            base = (_ROOT / base).resolve()
        tried.append(str(base))
        # Accept either the 13-CFS-ECY folder itself or a Data/ root above it.
        for probe in (base, base / "13-CFS-ECY", base / "Data" / "13-CFS-ECY"):
            if all((probe / f.filename).is_file() for f in FEEDS):
                return probe
    raise SourceError(
        "CFS/ECY CODECO corpus not found (need "
        + " + ".join(f.filename for f in FEEDS)
        + "). Looked in: " + ", ".join(tried)
        + ". Pass --corpus PATH or set JNPA_CORPUS_DIR.")


def parse_ts(raw: Any) -> Optional[dt.datetime]:
    """Parse a CODECO timestamp into an IST-aware datetime, or None."""
    if raw is None:
        return None
    if isinstance(raw, dt.datetime):
        return raw if raw.tzinfo else raw.replace(tzinfo=IST)
    s = str(raw).strip()
    if not s:
        return None
    for fmt in TS_FORMATS:
        try:
            return dt.datetime.strptime(s, fmt).replace(tzinfo=IST)
        except ValueError:
            continue
    return None


def parse_feed(path: Path, feed: Feed) -> Tuple[List[SourceEvent], List[dict]]:
    """Read one workbook into SourceEvents. Returns (events, rejected_rows).

    A row is rejected ONLY when it cannot be represented at all (no container
    number, unparseable timestamp, unrecognised mode). Rejections are reported
    and land in the DQ ledger — they are never silently dropped.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        try:
            header = [None if h is None else str(h).strip() for h in next(rows)]
        except StopIteration:
            raise SourceError(f"{path.name}: empty sheet")
        missing = [c for c in COLUMNS if c not in header]
        if missing:
            raise SourceError(
                f"{path.name}: expected columns {list(COLUMNS)}, got {header} "
                f"(missing {missing})")
        idx = {c: header.index(c) for c in COLUMNS}

        events: List[SourceEvent] = []
        rejected: List[dict] = []
        row_no = 0
        for values in rows:
            if not any(v is not None and str(v).strip() for v in values):
                continue                       # fully-blank spacer row
            row_no += 1

            def cell(col: str) -> Any:
                i = idx[col]
                return values[i] if i < len(values) else None

            raw_cn, raw_ts, raw_mode = (cell(c) for c in COLUMNS)
            cn = (str(raw_cn).strip().upper() if raw_cn is not None else "")
            ts = parse_ts(raw_ts)
            mode = (str(raw_mode).strip().upper() if raw_mode is not None else "")

            reason = None
            if not cn:
                reason = "missing_container"
            elif ts is None:
                reason = "bad_timestamp"
            elif mode not in ("IN", "OUT"):
                reason = "bad_mode"
            if reason:
                rejected.append({"file": feed.filename, "row": row_no,
                                 "reason": reason,
                                 "values": [raw_cn, raw_ts, raw_mode]})
                continue

            events.append(SourceEvent(
                container_no=cn,
                event_ts=ts,
                event_type=f"{feed.facility}_{mode}",
                location_type=feed.facility,
                direction="I" if mode == "IN" else "O",
                source_table=feed.source_table,
                iso_valid=bool(is_valid_container_no(cn)),
                raw_container=str(raw_cn).strip(),
                raw_timestamp=(raw_ts.isoformat() if isinstance(raw_ts, dt.datetime)
                               else str(raw_ts).strip()),
                raw_mode=str(raw_mode).strip(),
                source_file=feed.filename,
                source_row=row_no,
            ))
        return events, rejected
    finally:
        wb.close()


# ============================== source analysis ==============================
@dataclass
class Analysis:
    """Everything derived from the parsed source, with no DB involved."""

    per_feed: Dict[str, dict] = field(default_factory=dict)
    events: List[SourceEvent] = field(default_factory=list)
    rejected: List[dict] = field(default_factory=list)
    chains: Dict[str, dict] = field(default_factory=dict)     # complete chains only
    findings: List[dict] = field(default_factory=list)        # -> core.dq_issue

    @property
    def total_rows(self) -> int:
        return len(self.events) + len(self.rejected)


def _sample(values: Iterable[str], n: int = 5) -> str:
    picked = sorted(values)[:n]
    return ", ".join(picked) if picked else "—"


def analyse(events: Sequence[SourceEvent], rejected: Sequence[dict]) -> Analysis:
    """Derive per-feed inventory, the empty-container chains, and the DQ findings.

    Everything here is computed FROM THE SOURCE — no count is hard-coded, so a
    different corpus drop produces different (and still correct) numbers.
    """
    a = Analysis(events=list(events), rejected=list(rejected))

    by_type: Counter = Counter(e.event_type for e in events)
    for feed in FEEDS:
        feed_events = [e for e in events if e.source_file == feed.filename]
        feed_rejects = [r for r in rejected if r["file"] == feed.filename]
        span = [e.event_ts for e in feed_events]
        a.per_feed[feed.facility] = {
            "file": feed.filename,
            "expected_rows": feed.expected_rows,
            "parsed_rows": len(feed_events) + len(feed_rejects),
            "events": len(feed_events),
            "rejected": len(feed_rejects),
            "in_events": sum(1 for e in feed_events if e.direction == "I"),
            "out_events": sum(1 for e in feed_events if e.direction == "O"),
            "containers": len({e.container_no for e in feed_events}),
            "first_event_ts": min(span).isoformat() if span else None,
            "last_event_ts": max(span).isoformat() if span else None,
            "iso_invalid": sum(1 for e in feed_events if not e.iso_valid),
        }

    # --- per-container legs (the same aggregation mart.v_empty_container_trt does)
    legs: Dict[str, Dict[str, List[dt.datetime]]] = defaultdict(
        lambda: defaultdict(list))
    for e in events:
        legs[e.container_no][e.event_type].append(e.event_ts)

    for cn, by_leg in legs.items():
        ecy_out = min(by_leg["ECY_OUT"]) if by_leg.get("ECY_OUT") else None
        cfs_in = min(by_leg["CFS_IN"]) if by_leg.get("CFS_IN") else None
        cfs_out = max(by_leg["CFS_OUT"]) if by_leg.get("CFS_OUT") else None
        if ecy_out and cfs_in and cfs_out and cfs_in >= ecy_out and cfs_out >= cfs_in:
            a.chains[cn] = {
                "container_no": cn,
                "ecy_out_ts": ecy_out, "cfs_in_ts": cfs_in, "cfs_out_ts": cfs_out,
                "trt_min": round((cfs_in - ecy_out).total_seconds() / 60.0, 2),
                "dwell_min": round((cfs_out - cfs_in).total_seconds() / 60.0, 2),
                "cycle_min": round((cfs_out - ecy_out).total_seconds() / 60.0, 2),
            }

    ecy_out_containers = {e.container_no for e in events if e.event_type == "ECY_OUT"}
    ecy_in_containers = {e.container_no for e in events if e.event_type == "ECY_IN"}
    cfs_containers = {e.container_no for e in events
                      if e.event_type in ("CFS_IN", "CFS_OUT")}

    ecy = a.per_feed.get("ECY", {})
    cfs = a.per_feed.get("CFS", {})

    # ------------------------------------------------------------- findings
    # Grouped, one row per distinct anomaly — the granularity core.dq_issue is
    # already used at elsewhere in this schema. The per-container detail is
    # served live by GET /api/cfs-ecy/empty-trt/anomalies, so the ledger stays
    # readable without losing any of it.
    f: List[dict] = []

    # 1. The headline: the ECY log does not pair.
    if ecy and ecy["out_events"] != ecy["in_events"]:
        surplus = abs(ecy["out_events"] - ecy["in_events"])
        f.append({
            "feed": "ECY", "issue_type": "count_mismatch", "severity": "warn",
            "record_ref": ecy["file"],
            "description": (
                f"ECY CODECO log unpaired: {ecy['out_events']} OUT vs "
                f"{ecy['in_events']} IN events ({surplus} events cannot be paired). "
                f"All {ecy['events']} source rows imported verbatim — the surplus is "
                "NOT deleted, matched or completed with invented events."),
        })

    # 2. Why it does not pair: two date-disjoint blocks with no shared container.
    if ecy_out_containers and ecy_in_containers:
        overlap = ecy_out_containers & ecy_in_containers
        if not overlap:
            out_span = [e.event_ts for e in events if e.event_type == "ECY_OUT"]
            in_span = [e.event_ts for e in events if e.event_type == "ECY_IN"]
            f.append({
                "feed": "ECY", "issue_type": "disjoint_ranges", "severity": "warn",
                "record_ref": f"{ecy['file']}:ECY_OUT|ECY_IN",
                "description": (
                    f"ECY OUT block ({min(out_span):%Y-%m-%d} → {max(out_span):%Y-%m-%d}, "
                    f"{len(ecy_out_containers)} containers) and ECY IN block "
                    f"({min(in_span):%Y-%m-%d} → {max(in_span):%Y-%m-%d}, "
                    f"{len(ecy_in_containers)} containers) are date-disjoint and share "
                    "0 containers, so no container has both ECY legs. This is why the "
                    "customer's own mart.v_ecy_trt (ECY_IN → next ECY_OUT) returns no "
                    "rows; UC3-003 measures the ECY→CFS chain instead and leaves that "
                    "view untouched."),
            })

    # 3. ECY gate-outs that never reach a CFS.
    no_cfs = ecy_out_containers - cfs_containers
    if no_cfs:
        f.append({
            "feed": "ECY", "issue_type": "missing_key", "severity": "warn",
            "record_ref": f"{ecy['file']}:ECY_OUT",
            "description": (
                f"{len(no_cfs)} of {len(ecy_out_containers)} ECY gate-OUT containers "
                "have no CFS gate-IN anywhere in the corpus, so their empty-container "
                f"lifecycle cannot be closed. Excluded from the TRT KPI, retained in "
                f"full as PARTIAL chains. Examples: {_sample(no_cfs)}."),
        })

    # 4. ECY gate-ins with no preceding gate-out.
    orphan_ecy_in = ecy_in_containers - ecy_out_containers
    if orphan_ecy_in:
        f.append({
            "feed": "ECY", "issue_type": "missing_key", "severity": "warn",
            "record_ref": f"{ecy['file']}:ECY_IN",
            "description": (
                f"{len(orphan_ecy_in)} ECY gate-IN containers have no ECY gate-OUT and "
                "no CFS activity in the corpus — an empty returning to the yard whose "
                "outward leg is not in this drop. Imported verbatim and classified "
                f"ORPHAN. Examples: {_sample(orphan_ecy_in)}."),
        })

    # 5. CFS activity with no ECY origin.
    orphan_cfs = cfs_containers - ecy_out_containers
    if orphan_cfs:
        f.append({
            "feed": "CFS", "issue_type": "missing_key", "severity": "warn",
            "record_ref": f"{cfs['file']}:CFS_IN",
            "description": (
                f"{len(orphan_cfs)} of {len(cfs_containers)} CFS containers have CFS "
                "gate activity with no ECY gate-OUT in the corpus. Excluded from the "
                f"TRT KPI. Examples: {_sample(orphan_cfs)}."),
        })

    # 6. The CFS log, by contrast, is perfectly paired — worth flagging.
    if cfs and cfs["in_events"] == cfs["out_events"] and cfs["in_events"]:
        f.append({
            "feed": "CFS", "issue_type": "too_clean", "severity": "info",
            "record_ref": cfs["file"],
            "description": (
                f"CFS CODECO log is perfectly paired ({cfs['in_events']} IN / "
                f"{cfs['out_events']} OUT) — suspiciously clean next to the unpaired "
                "ECY log from the same drop."),
        })

    # 7. Exact-duplicate source rows: preserved, not collapsed.
    dupes = {k: c for k, c in Counter(e.key for e in events).items() if c > 1}
    if dupes:
        detail = "; ".join(
            f"{cn} {etype} {ts.astimezone(IST):%Y-%m-%d %H:%M} IST ×{n}"
            for (cn, ts, etype), n in sorted(dupes.items())[:5])
        f.append({
            "feed": "CFS", "issue_type": "duplicate", "severity": "warn",
            "record_ref": f"{cfs['file']}:duplicate-rows",
            "description": (
                f"{len(dupes)} gate event(s) appear more than once in the source and "
                "are stored with their source multiplicity — NOT de-duplicated. "
                "(Re-running the importer still inserts nothing: idempotency compares "
                f"per-event counts, not existence.) {detail}."),
        })

    # 8. More than one CFS gate-OUT for the same container.
    multi_out = sorted(cn for cn, by_leg in legs.items()
                       if len(by_leg.get("CFS_OUT", [])) > 1)
    if multi_out:
        f.append({
            "feed": "CFS", "issue_type": "duplicate", "severity": "warn",
            "record_ref": f"{cfs['file']}:CFS_OUT",
            "description": (
                f"{len(multi_out)} container(s) have more than one CFS gate-OUT; the "
                "chain uses the latest and every event is kept. Affected: "
                f"{_sample(multi_out)}."),
        })

    # 9. Container numbers failing the ISO-6346 check digit.
    iso_bad = sorted({e.container_no for e in events if not e.iso_valid})
    if iso_bad:
        f.append({
            "feed": "ECY", "issue_type": "bad_container_no", "severity": "info",
            "record_ref": "ISO-6346",
            "description": (
                f"{len(iso_bad)} distinct container number(s) fail the ISO-6346 check "
                f"digit. Stored verbatim, never corrected. Examples: {_sample(iso_bad)}."),
        })

    # 10. Rows that could not be represented at all.
    if rejected:
        by_reason = Counter(r["reason"] for r in rejected)
        f.append({
            "feed": "ECY", "issue_type": "parse_error", "severity": "error",
            "record_ref": "unparseable-rows",
            "description": (
                f"{len(rejected)} source row(s) could not be mapped to a gate event "
                f"and were not imported: {dict(by_reason)}."),
        })

    a.findings = f
    return a


def trt_summary(analysis: Analysis) -> dict:
    """KPI-3 roll-up over the COMPLETE chains, using the project's KPI engine."""
    from jnpa_shared import kpi as kpi_engine

    chains = list(analysis.chains.values())
    trts = sorted(c["trt_min"] for c in chains)
    if not trts:
        return {"valid_containers": 0, "kpi": None}
    # The project's own aggregation helper: mean of the ECD-pickup -> gate-in
    # samples, in minutes. Fed with seconds, exactly as its signature asks.
    value = kpi_engine.trt_empty_ecd_min([t * 60.0 for t in trts])
    mid = len(trts) // 2
    median = trts[mid] if len(trts) % 2 else (trts[mid - 1] + trts[mid]) / 2.0
    return {
        "valid_containers": len(chains),
        "avg_trt_min": round(value, 2),
        "median_trt_min": round(median, 2),
        "min_trt_min": trts[0],
        "max_trt_min": trts[-1],
        "kpi": kpi_engine.compute_kpi("trt_empty_ecd", round(value, 2),
                                      source="live", n=len(chains)).to_dict(),
    }


# ================================== database =================================
_UPSERT_FILE = """
INSERT INTO core.ingest_file (path, source_system, file_format, row_count, notes)
VALUES (:path, :source_system, 'xlsx', :row_count, :notes)
ON CONFLICT (path) DO UPDATE
    SET row_count = EXCLUDED.row_count,
        source_system = EXCLUDED.source_system
RETURNING file_id
"""

_EXISTING = """
SELECT container_no, event_ts, event_type, count(*) AS n
FROM core.container_event
WHERE source_table = ANY(:tables)
GROUP BY container_no, event_ts, event_type
"""

_INSERT_EVENT = """
INSERT INTO core.container_event
    (container_no, event_ts, event_type, location_type, direction,
     source_table, source_file, details)
VALUES
    (:container_no, :event_ts, :event_type, :location_type, :direction,
     :source_table, :source_file, CAST(:details AS jsonb))
"""

# Only this importer's own findings are cleared before they are re-written, so a
# re-run refreshes UC3-003's ledger entries and touches no other module's rows.
_CLEAR_DQ = """
DELETE FROM core.dq_issue
WHERE source_table = :source_table AND file_id = ANY(:file_ids)
"""

_INSERT_DQ = """
INSERT INTO core.dq_issue
    (file_id, source_table, record_ref, issue_type, severity, description)
VALUES (:file_id, :source_table, :record_ref, :issue_type, :severity, :description)
"""


async def run_import(analysis: Analysis, dsn: str) -> dict:
    """Register the source files, insert the missing events, refresh the DQ ledger.

    One transaction. Returns per-stage counters so a caller (or the CLI) can show
    first-run vs re-run evidence.
    """
    from sqlalchemy import text

    from jnpa_shared.db import get_engine

    stats: Dict[str, Any] = {"files": 0, "events_source": len(analysis.events),
                             "events_inserted": 0, "events_already_present": 0,
                             "dq_cleared": 0, "dq_written": 0}

    engine = get_engine(dsn)
    async with engine.begin() as conn:
        # --- 1. source files (natural key = path, so a re-run reuses the row)
        file_ids: Dict[str, int] = {}
        for feed in FEEDS:
            info = analysis.per_feed.get(feed.facility, {})
            row = (await conn.execute(text(_UPSERT_FILE), {
                "path": feed.ingest_path,
                "source_system": INGEST_SOURCE_SYSTEM,
                "row_count": info.get("parsed_rows"),
                "notes": None,
            })).mappings().first()
            file_ids[feed.facility] = int(row["file_id"])
            stats["files"] += 1

        # --- 2. events, multiplicity-aware
        present: Dict[Tuple[str, dt.datetime, str], int] = {}
        for row in (await conn.execute(
                text(_EXISTING),
                {"tables": [f.source_table for f in FEEDS]})).mappings():
            present[(row["container_no"], row["event_ts"], row["event_type"])] = \
                int(row["n"])

        wanted: Dict[Tuple[str, dt.datetime, str], List[SourceEvent]] = defaultdict(list)
        for e in analysis.events:
            wanted[e.key].append(e)

        pending: List[dict] = []
        for key, group in wanted.items():
            have = present.get(key, 0)
            stats["events_already_present"] += min(have, len(group))
            for e in group[have:]:
                pending.append({
                    "container_no": e.container_no,
                    "event_ts": e.event_ts,
                    "event_type": e.event_type,
                    "location_type": e.location_type,
                    "direction": e.direction,
                    "source_table": e.source_table,
                    "source_file": file_ids[e.location_type],
                    # Source values kept verbatim alongside the typed columns, so
                    # the original spreadsheet cell is always recoverable.
                    "details": json.dumps({
                        "source_file": e.source_file,
                        "source_row": e.source_row,
                        "source_container": e.raw_container,
                        "source_timestamp": e.raw_timestamp,
                        "source_mode": e.raw_mode,
                        "iso6346_valid": e.iso_valid,
                        "feed": "CODECO",
                    }, ensure_ascii=False),
                })

        for i in range(0, len(pending), BATCH):
            chunk = pending[i:i + BATCH]
            if chunk:
                await conn.execute(text(_INSERT_EVENT), chunk)
                stats["events_inserted"] += len(chunk)

        # --- 3. data-quality ledger
        res = await conn.execute(text(_CLEAR_DQ), {
            "source_table": DQ_SOURCE_TABLE,
            "file_ids": list(file_ids.values())})
        stats["dq_cleared"] = res.rowcount or 0
        for finding in analysis.findings:
            await conn.execute(text(_INSERT_DQ), {
                "file_id": file_ids.get(finding["feed"]),
                "source_table": DQ_SOURCE_TABLE,
                "record_ref": finding["record_ref"],
                "issue_type": finding["issue_type"],
                "severity": finding["severity"],
                "description": finding["description"],
            })
            stats["dq_written"] += 1

    stats["file_ids"] = file_ids
    return stats


# ==================================== CLI ====================================
def build(corpus: Path) -> Analysis:
    events: List[SourceEvent] = []
    rejected: List[dict] = []
    for feed in FEEDS:
        e, r = parse_feed(corpus / feed.filename, feed)
        events.extend(e)
        rejected.extend(r)
    return analyse(events, rejected)


def report(analysis: Analysis, corpus: Path, stats: Optional[dict]) -> None:
    trt = trt_summary(analysis)
    line = "=" * 74
    print("\n" + line)
    print("UC3-003 — CFS/ECY CODECO GATE EVENTS" + ("" if stats else "   [DRY-RUN]"))
    print(line)
    print(f"  corpus: {corpus}")
    for facility, info in analysis.per_feed.items():
        flag = "" if info["parsed_rows"] == info["expected_rows"] else \
            f"  (brief expects {info['expected_rows']})"
        print(f"  {facility} {info['file']:18} rows={info['parsed_rows']:5}{flag}"
              f"  IN={info['in_events']:4}  OUT={info['out_events']:4}"
              f"  containers={info['containers']:4}  rejected={info['rejected']}")
    print(f"  total source rows              : {analysis.total_rows}")
    print(f"  events mapped                  : {len(analysis.events)}")
    print("-" * 74)
    print(f"  COMPLETE chains (ECY-Out → CFS-In → CFS-Out): {trt['valid_containers']}")
    if trt["kpi"]:
        k = trt["kpi"]
        print(f"  TRT (ECD pickup → CFS gate-in) : avg {trt['avg_trt_min']} min   "
              f"median {trt['median_trt_min']}   min {trt['min_trt_min']}   "
              f"max {trt['max_trt_min']}")
        print(f"  KPI {k['key']}: target {k['target']} min · baseline "
              f"{k['baseline']} min · vs baseline {k['deltaPct']:+.1f}% · "
              f"onTarget={k['onTarget']}")
    print("-" * 74)
    print(f"  DATA QUALITY (core.dq_issue) — {len(analysis.findings)} grouped findings")
    for f in analysis.findings:
        print(f"   [{f['severity']:5}] {f['issue_type']:17} {f['record_ref']}")
        print(f"           {f['description']}")
    if stats:
        print("-" * 74)
        print(f"  ingest_file rows registered    : {stats['files']} {stats['file_ids']}")
        print(f"  events inserted (new)          : {stats['events_inserted']}")
        print(f"  events already present         : {stats['events_already_present']}")
        print(f"  dq rows replaced / written     : {stats['dq_cleared']} / "
              f"{stats['dq_written']}")
    print(line + "\n")


def main(argv: Optional[Sequence[str]] = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--corpus", default=None,
                    help="directory holding ECY-CODECO.xlsx + CFS-CODECO.xlsx "
                         "(default: discovered, or $JNPA_CORPUS_DIR)")
    ap.add_argument("--dsn", default=os.environ.get("POSTGRES_DSN", "") or None,
                    help="SQLAlchemy asyncpg DSN (default: $POSTGRES_DSN)")
    ap.add_argument("--dry-run", action="store_true",
                    help="parse, analyse and report; write nothing")
    ap.add_argument("--json", action="store_true",
                    help="emit the analysis as JSON instead of the text report")
    args = ap.parse_args(argv)

    try:
        corpus = find_corpus(args.corpus)
        analysis = build(corpus)
    except SourceError as exc:
        print(f"\nSOURCE ERROR: {exc}", file=sys.stderr)
        return 2

    stats = None
    if not args.dry_run:
        if not args.dsn:
            print("\nERROR: no DSN — pass --dsn or set POSTGRES_DSN.", file=sys.stderr)
            return 2
        stats = asyncio.run(run_import(analysis, args.dsn))

    if args.json:
        print(json.dumps({
            "corpus": str(corpus),
            "per_feed": analysis.per_feed,
            "total_rows": analysis.total_rows,
            "trt": trt_summary(analysis),
            "findings": analysis.findings,
            "import": stats,
        }, indent=2, default=str))
    else:
        report(analysis, corpus, stats)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
