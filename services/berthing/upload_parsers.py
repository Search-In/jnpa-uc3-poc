"""Berthing UPLOAD parsers — template, byte readers, validation & column mapping.

The reusable Data-Upload sub-module for module 7. Mirrors
:mod:`services.cfs_ecy.upload_parsers`: pure functions that turn an uploaded
CSV/XLS/XLSX byte payload into a validated, mapped record set plus a preview and
user-friendly errors — WITHOUT touching the DB. The import step hands the valid
records to :class:`services.berthing.repository.BerthingRepository.persist` (the SAME
normalised model the PDF importer produces).

Column mapping is ALIAS-DRIVEN (header normalised, then matched against an alias
table), so "Vessel", "Vessel Name", "VESSEL_NAME" all map to one field. Required per
the real corpus: Terminal, Vessel Name, Voyage Number (imo_number is absent in every
source, NSFT has no berth, and ETA only appears in the Expected section — so those
stay optional).
"""
from __future__ import annotations

import csv
import datetime as _dt
import io
import re
from typing import Any, Optional

IST = _dt.timezone(_dt.timedelta(hours=5, minutes=30))

TERMINALS = ("APMT", "BMCT", "NSFT", "NSICT", "NSIGT")
STATUSES = ("EXPECTED", "ARRIVED", "BERTH_ASSIGNED", "BERTHING_STARTED",
            "CARGO_OPERATION", "COMPLETED", "DEPARTED")

# Terminal-name synonyms → canonical code.
_TERMINAL_ALIASES = {
    "apmt": "APMT", "apm": "APMT", "apmterminals": "APMT", "apmterminalsmumbai": "APMT",
    "bmct": "BMCT", "bmctpsa": "BMCT", "psa": "BMCT",
    "nsft": "NSFT", "nhavashevafreeportterminal": "NSFT", "freeport": "NSFT",
    "nsict": "NSICT", "nsictdpworld": "NSICT",
    "nsigt": "NSIGT", "nsigtdpworld": "NSIGT",
}

# canonical field -> accepted NORMALISED header names (first present, non-empty wins).
ALIASES: dict[str, tuple[str, ...]] = {
    "terminal": ("terminal", "terminalname", "terminalcode", "port", "facility"),
    "vessel_name": ("vesselname", "vessel", "vesselnm", "shipname", "ship", "name"),
    "imo_number": ("imonumber", "imono", "imo", "imonbr"),
    "voyage_number": ("voyagenumber", "voyageno", "voyage", "via", "viano", "vianumber",
                      "rotation", "rotationno", "rotationnumber"),
    "shipping_line": ("shippingline", "line", "linecode", "carrier", "operator", "voa"),
    "berth_number": ("berthnumber", "berthno", "berth", "berthcode"),
    "eta": ("eta", "expectedarrival", "etadatetime", "expectedtimeofarrival"),
    "ata": ("ata", "actualarrival", "alongside", "berthed", "arrivaltime", "arrival"),
    "berthing_time": ("berthingtime", "berthing", "madefast", "firstline"),
    "departure_time": ("departuretime", "departure", "atd", "sailed", "sailingtime",
                       "etd", "sailtime"),
    "cargo_operation_start": ("cargooperationstart", "opscommenced", "operationstart",
                              "opsstart", "cargostart", "commenced"),
    "cargo_operation_end": ("cargooperationend", "opscompleted", "operationend", "opsend",
                            "cargoend", "completed", "etc", "atc"),
    "status": ("status", "vesselstatus", "state", "stage"),
}

# canonical label shown to the user -> the alias tuple that satisfies it. Terminal is
# NOT a required COLUMN — it may come from the upload's terminal selector (a per-row
# Terminal column overrides it); it is enforced as a VALUE per row (invalid_terminal).
_REQUIRED = {
    "Vessel Name": ALIASES["vessel_name"],
    "Voyage Number": ALIASES["voyage_number"],
}

_TEMPLATE_COLS = ["Terminal", "Vessel Name", "IMO Number", "Voyage Number",
                  "Shipping Line", "Berth Number", "ETA", "ATA", "Berthing Time",
                  "Departure Time", "Cargo Operation Start", "Cargo Operation End",
                  "Status"]
_TEMPLATE_EXAMPLE = ["NSICT", "MAERSK SENTOSA", "", "S0488", "MSK", "CB05",
                     "05/06/2026 16:00", "05/06/2026 09:00", "05/06/2026 09:00", "",
                     "05/06/2026 09:40", "", "CARGO_OPERATION"]

_TS_FORMATS = (
    "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%d-%m-%Y %H:%M:%S",
    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d",
)


def norm_header(name: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(name).strip().lower())


def clean(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    v = str(raw).strip()
    return v or None


def terminal_ok(value: Any) -> Optional[str]:
    v = norm_header(value)
    return _TERMINAL_ALIASES.get(v) or (value.strip().upper()
                                        if clean(value) and value.strip().upper() in TERMINALS
                                        else None)


def status_ok(value: Any) -> Optional[str]:
    v = (clean(value) or "").upper().replace(" ", "_").replace("-", "_")
    return v if v in STATUSES else None


def parse_ts(raw: Any) -> Optional[_dt.datetime]:
    if raw is None:
        return None
    if isinstance(raw, _dt.datetime):
        return raw if raw.tzinfo else raw.replace(tzinfo=IST)
    if isinstance(raw, _dt.date):
        return _dt.datetime(raw.year, raw.month, raw.day, tzinfo=IST)
    s = clean(raw)
    if not s:
        return None
    for fmt in _TS_FORMATS:
        try:
            return _dt.datetime.strptime(s, fmt).replace(tzinfo=IST)
        except ValueError:
            continue
    try:
        d = _dt.datetime.fromisoformat(s.replace("Z", "+00:00"))
        return d if d.tzinfo else d.replace(tzinfo=IST)
    except ValueError:
        return None


def template_csv() -> str:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(_TEMPLATE_COLS)
    guide = ["# REQUIRED: Terminal, Vessel Name, Voyage Number. Terminal ∈ "
             "APMT/BMCT/NSFT/NSICT/NSIGT. Timestamps DD/MM/YYYY HH:MM (IST). "
             "Status ∈ EXPECTED/ARRIVED/BERTH_ASSIGNED/BERTHING_STARTED/"
             "CARGO_OPERATION/COMPLETED/DEPARTED (derived if blank). Delete this "
             "line and the example before uploading."] + [""] * (len(_TEMPLATE_COLS) - 1)
    w.writerow(guide)
    w.writerow(_TEMPLATE_EXAMPLE)
    return buf.getvalue()


# ---------------------------------------------------------------- byte readers
def read_rows_from_bytes(content: bytes, filename: str) -> tuple[list[str], list[dict[str, Any]]]:
    """Return (header, rows) from a CSV / XLS / XLSX byte payload. Raises ValueError on
    an unreadable/empty file or unsupported extension (PDF is handled elsewhere)."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            it = ws.iter_rows(values_only=True)
            try:
                header = [str(c).strip() if c is not None else "" for c in next(it)]
            except StopIteration:
                raise ValueError("empty_file")
            rows = []
            for values in it:
                if not any(v not in (None, "") for v in values):
                    continue
                rows.append({header[i]: (values[i] if i < len(values) else None)
                             for i in range(len(header))})
        finally:
            wb.close()
        return header, rows
    if name.endswith(".xls"):
        import xlrd
        book = xlrd.open_workbook(file_contents=content)
        sh = book.sheet_by_index(0)
        if sh.nrows == 0:
            raise ValueError("empty_file")
        header = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]
        rows = []
        for r in range(1, sh.nrows):
            values = [sh.cell_value(r, c) for c in range(sh.ncols)]
            if not any(str(v).strip() for v in values):
                continue
            rows.append({header[i]: values[i] for i in range(len(header))})
        return header, rows
    if name.endswith((".csv", ".txt")) or name == "":
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        all_rows = [r for r in reader if any((c or "").strip() for c in r)]
        if not all_rows:
            raise ValueError("empty_file")
        header = [c.strip() for c in all_rows[0]]
        rows = []
        for r in all_rows[1:]:
            if r and str(r[0]).strip().startswith("#"):     # skip template guidance line
                continue
            rows.append({header[i]: (r[i] if i < len(r) else None) for i in range(len(header))})
        return header, rows
    raise ValueError("unsupported_format")


# ---------------------------------------------------------------- ParseResult
class ParseResult:
    def __init__(self) -> None:
        self.errors: list[dict[str, Any]] = []
        self.warnings: list[dict[str, Any]] = []
        self.records: list[dict[str, Any]] = []
        self.preview: list[dict[str, Any]] = []
        self.row_count = 0
        self.invalid_count = 0
        self.duplicate_count = 0
        self.rejected = False

    def err(self, row, col, code, detail, raw=None):
        self.errors.append({"row_number": row, "column_name": col, "error_code": code,
                            "error_detail": detail, "raw_value": (None if raw is None else str(raw))})

    def warn(self, row, col, code, detail):
        self.warnings.append({"row_number": row, "column_name": col, "error_code": code,
                              "error_detail": detail})


def check_required_columns(res: ParseResult, header: list[str]) -> bool:
    hset = {norm_header(h) for h in header if norm_header(h)}
    missing = [label for label, aliases in _REQUIRED.items()
               if not any(a in hset for a in aliases)]
    if missing:
        for label in missing:
            res.err(None, label, "missing_column",
                    f"{label} column not found. Please download the latest template.")
        res.rejected = True
        return False
    return True


def _pick(row_norm: dict[str, Any], canonical: str) -> Optional[str]:
    for src in ALIASES.get(canonical, ()):
        if src in row_norm:
            v = clean(row_norm[src])
            if v is not None:
                return v
    return None


_TS_FIELDS = ("eta", "ata", "berthing_time", "departure_time",
              "cargo_operation_start", "cargo_operation_end")


def parse(header: list[str], rows: list[dict[str, Any]], *,
          terminal: Optional[str] = None, source_file: Optional[str] = None) -> ParseResult:
    """Validate + map berthing rows for one uploaded file. ``terminal`` is the selector
    default used when a row carries no Terminal column value."""
    res = ParseResult()
    res.row_count = len(rows)
    if not check_required_columns(res, header):
        return res

    seen: set[tuple[str, str, str]] = set()
    for i, raw in enumerate(rows, start=1):
        row_norm = {norm_header(k): v for k, v in raw.items() if norm_header(k)}

        vessel = _pick(row_norm, "vessel_name")
        if not vessel:
            res.err(i, "Vessel Name", "empty_required", "vessel name is empty")
            res.invalid_count += 1
            continue
        vessel = vessel.upper()

        voyage = _pick(row_norm, "voyage_number")
        if not voyage:
            res.err(i, "Voyage Number", "empty_required", "voyage number is empty")
            res.invalid_count += 1
            continue
        voyage = voyage.upper()

        raw_term = _pick(row_norm, "terminal")
        term = terminal_ok(raw_term) if raw_term is not None else terminal_ok(terminal)
        if term is None:
            res.err(i, "Terminal", "invalid_terminal",
                    f"terminal '{raw_term or terminal}' is not one of "
                    "APMT/BMCT/NSFT/NSICT/NSIGT", raw_term or terminal)
            res.invalid_count += 1
            continue

        rec: dict[str, Any] = {
            "terminal": term, "vessel_name": vessel, "voyage_number": voyage,
            "imo_number": _pick(row_norm, "imo_number"),
            "shipping_line": (_pick(row_norm, "shipping_line") or None),
            "berth_number": (_pick(row_norm, "berth_number") or None),
            "source_file": source_file,
        }
        if rec["shipping_line"]:
            rec["shipping_line"] = rec["shipping_line"].upper()
        if rec["berth_number"]:
            rec["berth_number"] = rec["berth_number"].upper()

        bad_ts = False
        for f in _TS_FIELDS:
            raw_v = _pick(row_norm, f)
            if raw_v is None:
                rec[f] = None
                continue
            v = parse_ts(raw_v)
            if v is None:
                res.err(i, f, "invalid_timestamp",
                        f"'{raw_v}' is not a recognised date/time (expected DD/MM/YYYY HH:MM)",
                        raw_v)
                bad_ts = True
                break
            rec[f] = v
        if bad_ts:
            res.invalid_count += 1
            continue

        raw_status = _pick(row_norm, "status")
        if raw_status is not None:
            st = status_ok(raw_status)
            if st is None:
                res.err(i, "Status", "invalid_status",
                        f"status '{raw_status}' is not a recognised lifecycle state", raw_status)
                res.invalid_count += 1
                continue
            rec["status"] = st
        else:
            rec["status"] = _derive_status(rec)

        key = (rec["terminal"], rec["voyage_number"], rec["vessel_name"])
        if key in seen:
            res.duplicate_count += 1
            res.warn(i, "Voyage Number", "duplicate_in_file",
                     f"{vessel} / {voyage} ({term}) already appears earlier in this file (skipped)")
            continue
        seen.add(key)
        res.records.append(rec)

    res.preview = [{
        "Terminal": r["terminal"], "Vessel": r["vessel_name"], "Voyage": r["voyage_number"],
        "Berth": r.get("berth_number") or "—", "Status": r["status"],
        "ETA": r["eta"].strftime("%d/%m/%Y %H:%M") if r.get("eta") else "—",
        "ATA": r["ata"].strftime("%d/%m/%Y %H:%M") if r.get("ata") else "—",
    } for r in res.records[:20]]
    return res


def _derive_status(rec: dict[str, Any]) -> str:
    """Infer a lifecycle status from whichever timestamps are present."""
    if rec.get("departure_time"):
        return "DEPARTED"
    if rec.get("cargo_operation_end"):
        return "COMPLETED"
    if rec.get("cargo_operation_start"):
        return "CARGO_OPERATION"
    if rec.get("berthing_time") or rec.get("ata"):
        return "BERTH_ASSIGNED" if rec.get("berth_number") else "ARRIVED"
    return "EXPECTED"
