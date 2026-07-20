"""CFS-ECY UPLOAD parsers — template, byte readers, validation & column mapping.

The reusable Data-Upload sub-module (module 13). Mirrors
:mod:`services.shipping_lines.upload_parsers`: pure functions that turn an uploaded
CSV/XLS/XLSX byte payload into a validated, mapped record set plus a preview and
user-friendly errors — WITHOUT touching the DB. The import step then hands the valid
records to :class:`services.cfs_ecy.repository.CfsEcyRepository.persist` (the SAME
target table jnpa.cfs_ecy_movements + its (facility_type, container_number, event_ts,
mode) UNIQUE key → idempotent, duplicate-safe).

Column mapping is ALIAS-DRIVEN (header is normalised, then matched against an alias
table), so "Container No" / "Container Number" / "CNTR_NO" / "Container_No" all map to
one field. The CODECO feed carries three columns — Container Number, Timestamp, Mode —
and the facility (CFS / ECY) is NOT a column in the JNPA files, so it is supplied by
the upload's facility selector; an optional Facility column is honoured per-row if the
file happens to carry one (selector is the fallback default).
"""
from __future__ import annotations

import csv
import datetime as _dt
import io
import re
from typing import Any, Optional

from jnpa_shared.iso6346 import is_valid_container_no

# JNPA operates in IST. The CODECO timestamps carry no timezone; we stamp them as
# Asia/Kolkata (UTC+5:30) so the timestamptz column stores the correct instant —
# identical to scripts/import_cfs_ecy_codeco.py.
IST = _dt.timezone(_dt.timedelta(hours=5, minutes=30))

FACILITIES = ("CFS", "ECY")

# ---------------------------------------------------------------- column aliases
# canonical field -> accepted NORMALISED header names (see norm_header). First
# present, non-empty match wins.
ALIASES: dict[str, tuple[str, ...]] = {
    "container_number": (
        "containerno", "containernumber", "containernbr", "container", "cntrno",
        "cntrnumber", "cntr", "containerid", "equipmentno", "equipmentnumber",
        "unitno", "boxno", "containerno1",
    ),
    "event_ts": (
        "timestamp", "eventts", "eventtime", "eventtimestamp", "eventdatetime",
        "gatedatetime", "gatetimestamp", "gatetime", "gatedate", "datetime",
        "date", "datetimeist", "movementtime", "movementdatetime", "ts", "time",
    ),
    "mode": (
        "mode", "movement", "movementtype", "inout", "innout", "direction",
        "gatemode", "gatemovement", "type",
    ),
    "facility_type": (
        "facility", "facilitytype", "facilitycode", "location", "yard", "cfsecy",
        "cfsecytype", "site",
    ),
}

# canonical label shown to the user -> the alias tuple that satisfies it (facility is
# NOT required — it comes from the selector).
_REQUIRED = {
    "Container Number": ALIASES["container_number"],
    "Timestamp": ALIASES["event_ts"],
    "Mode": ALIASES["mode"],
}

# ---------------------------------------------------------------- template
_TEMPLATE_COLS = ["Container Number", "Timestamp", "Mode", "Facility"]
_TEMPLATE_EXAMPLE = ["ONEU2122848", "01/07/2026 14:00", "In", "CFS"]


def facility_ok(value: Optional[str]) -> Optional[str]:
    v = (value or "").strip().upper()
    return v if v in FACILITIES else None


def norm_header(name: Any) -> str:
    """Lowercase and drop everything that is not a letter or digit so header
    variations ('Container No', 'CNTR_NO', 'Container_Number') collapse to one key."""
    return re.sub(r"[^a-z0-9]", "", str(name).strip().lower())


def clean(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    v = str(raw).strip()
    return v or None


def norm_container(raw: Any) -> Optional[str]:
    v = clean(raw)
    return v.upper().replace(" ", "") if v else None


def norm_mode(raw: Any) -> Optional[str]:
    """In -> IN, Out -> OUT (case-insensitive). Also tolerant of gate-in/gate-out,
    i/o. Anything else -> None (invalid)."""
    v = (clean(raw) or "").upper()
    if v in ("IN", "GATEIN", "GATE-IN", "GATE IN", "I", "GI"):
        return "IN"
    if v in ("OUT", "GATEOUT", "GATE-OUT", "GATE OUT", "O", "GO"):
        return "OUT"
    return None


_TS_FORMATS = (
    "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%d-%m-%Y %H:%M:%S",
    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d",
)


def parse_ts(raw: Any) -> Optional[_dt.datetime]:
    """Parse a CODECO timestamp into an IST-stamped aware datetime. Accepts an Excel
    native datetime, the DD/MM/YYYY HH:MM family, and ISO — else None (invalid)."""
    if raw is None:
        return None
    if isinstance(raw, _dt.datetime):
        return raw if raw.tzinfo else raw.replace(tzinfo=IST)
    if isinstance(raw, _dt.date):
        return _dt.datetime(raw.year, raw.month, raw.day, tzinfo=IST)
    s = clean(raw)
    if not s:
        return None
    for fmt in _TS_FORMATS:
        try:
            return _dt.datetime.strptime(s, fmt).replace(tzinfo=IST)
        except ValueError:
            continue
    try:
        d = _dt.datetime.fromisoformat(s.replace("Z", "+00:00"))
        return d if d.tzinfo else d.replace(tzinfo=IST)
    except ValueError:
        return None


def template_csv() -> str:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(_TEMPLATE_COLS)
    w.writerow([
        "# REQUIRED: Container Number, Timestamp, Mode | Facility is OPTIONAL "
        "(the CFS/ECY selector is used when this column is absent). "
        "Timestamp format DD/MM/YYYY HH:MM (IST). Mode = In / Out. "
        "Column names are flexible (e.g. 'Container No' / 'CNTR_NO' also work). "
        "Delete this line and the example row before uploading.",
        "", "", "",
    ])
    w.writerow(_TEMPLATE_EXAMPLE)
    return buf.getvalue()


# ---------------------------------------------------------------- byte readers
def read_rows_from_bytes(content: bytes, filename: str) -> tuple[list[str], list[dict[str, Any]]]:
    """Return (header, rows) from a CSV / XLS / XLSX byte payload. Raises ValueError on
    an unreadable/empty file or unsupported extension."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            it = ws.iter_rows(values_only=True)
            try:
                header = [str(c).strip() if c is not None else "" for c in next(it)]
            except StopIteration:
                raise ValueError("empty_file")
            rows = []
            for values in it:
                if not any(v not in (None, "") for v in values):
                    continue
                rows.append({header[i]: (values[i] if i < len(values) else None)
                             for i in range(len(header))})
        finally:
            wb.close()
        return header, rows
    if name.endswith(".xls"):
        import xlrd
        book = xlrd.open_workbook(file_contents=content)
        sh = book.sheet_by_index(0)
        if sh.nrows == 0:
            raise ValueError("empty_file")
        header = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]
        rows = []
        for r in range(1, sh.nrows):
            values = [sh.cell_value(r, c) for c in range(sh.ncols)]
            if not any(str(v).strip() for v in values):
                continue
            rows.append({header[i]: values[i] for i in range(len(header))})
        return header, rows
    if name.endswith((".csv", ".txt")) or name == "":
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        all_rows = [r for r in reader if any((c or "").strip() for c in r)]
        if not all_rows:
            raise ValueError("empty_file")
        header = [c.strip() for c in all_rows[0]]
        rows = []
        for r in all_rows[1:]:
            if r and str(r[0]).strip().startswith("#"):   # skip template guidance lines
                continue
            rows.append({header[i]: (r[i] if i < len(r) else None) for i in range(len(header))})
        return header, rows
    raise ValueError("unsupported_format")


# ---------------------------------------------------------------- ParseResult
class ParseResult:
    def __init__(self) -> None:
        self.errors: list[dict[str, Any]] = []
        self.warnings: list[dict[str, Any]] = []
        self.records: list[dict[str, Any]] = []   # valid, mapped canonical rows
        self.preview: list[dict[str, Any]] = []
        self.row_count = 0
        self.invalid_count = 0
        self.duplicate_count = 0
        self.rejected = False                       # structural failure (wrong template)

    def err(self, row: Optional[int], col: Optional[str], code: str, detail: str, raw: Any = None):
        self.errors.append({"row_number": row, "column_name": col, "error_code": code,
                            "error_detail": detail, "raw_value": (None if raw is None else str(raw))})

    def warn(self, row: Optional[int], col: Optional[str], code: str, detail: str):
        self.warnings.append({"row_number": row, "column_name": col, "error_code": code,
                              "error_detail": detail})


def check_required_columns(res: ParseResult, header: list[str]) -> bool:
    """Alias-aware required-column check. Missing → user-friendly error + rejected."""
    hset = {norm_header(h) for h in header if norm_header(h)}
    missing = []
    for label, aliases in _REQUIRED.items():
        if not any(a in hset for a in aliases):
            missing.append(label)
    if missing:
        for label in missing:
            res.err(None, label, "missing_column",
                    f"{label} column not found. Please download the latest template.")
        res.rejected = True
        return False
    return True


def _pick(row_norm: dict[str, Any], canonical: str) -> Optional[str]:
    for src in ALIASES.get(canonical, ()):
        if src in row_norm:
            v = clean(row_norm[src])
            if v is not None:
                return v
    return None


# ---------------------------------------------------------------- main parse
def parse(header: list[str], rows: list[dict[str, Any]], *, facility: str,
          source_file: Optional[str] = None) -> ParseResult:
    """Validate + map CODECO rows for one uploaded file.

    ``facility`` is the selector value (CFS / ECY) used when a row carries no Facility
    column. A per-row Facility column, if present and valid, overrides the selector."""
    res = ParseResult()
    res.row_count = len(rows)
    if not check_required_columns(res, header):
        return res

    seen: set[tuple[str, str, Any, str]] = set()
    for i, raw in enumerate(rows, start=1):
        row_norm = {norm_header(k): v for k, v in raw.items() if norm_header(k)}

        cn = norm_container(_pick(row_norm, "container_number"))
        if not cn:
            res.err(i, "Container Number", "empty_container", "container number is empty")
            res.invalid_count += 1
            continue

        # facility: per-row column wins, else the selector default.
        raw_fac = _pick(row_norm, "facility_type")
        if raw_fac is not None:
            fac = facility_ok(raw_fac)
            if fac is None:
                res.err(i, "Facility", "invalid_facility",
                        f"facility '{raw_fac}' is not CFS or ECY", raw_fac)
                res.invalid_count += 1
                continue
        else:
            fac = facility

        raw_ts = _pick(row_norm, "event_ts")
        if raw_ts is None:
            res.err(i, "Timestamp", "empty_required", "timestamp is empty")
            res.invalid_count += 1
            continue
        ts = parse_ts(raw_ts)
        if ts is None:
            res.err(i, "Timestamp", "invalid_timestamp",
                    f"timestamp '{raw_ts}' is not a recognised date/time "
                    "(expected DD/MM/YYYY HH:MM)", raw_ts)
            res.invalid_count += 1
            continue

        raw_mode = _pick(row_norm, "mode")
        mode = norm_mode(raw_mode)
        if mode is None:
            res.err(i, "Mode", "invalid_mode",
                    f"mode '{raw_mode}' is not IN or OUT", raw_mode)
            res.invalid_count += 1
            continue

        iso_valid = bool(is_valid_container_no(cn))
        if not iso_valid:
            res.warn(i, "Container Number", "container_iso6346_invalid",
                     f"{cn} fails the ISO-6346 check digit (imported, flagged)")

        # duplicate within THIS file (the same key the DB unique constraint enforces)
        key = (fac, cn, ts, mode)
        if key in seen:
            res.duplicate_count += 1
            res.warn(i, "Container Number", "duplicate_in_file",
                     f"{cn} {mode} @ {ts:%d/%m/%Y %H:%M} ({fac}) already appears earlier "
                     "in this file (skipped)")
            continue
        seen.add(key)

        res.records.append({
            "facility_type": fac,
            "container_number": cn,
            "iso_valid": iso_valid,
            "event_ts": ts,
            "mode": mode,
            "source": "UPLOAD",
            "source_file": source_file,
        })

    res.preview = [{
        "Facility": r["facility_type"],
        "Container": r["container_number"],
        "Mode": r["mode"],
        "Timestamp": r["event_ts"].strftime("%d/%m/%Y %H:%M"),
        "ISO": "valid" if r["iso_valid"] else "invalid",
    } for r in res.records[:20]]
    return res
