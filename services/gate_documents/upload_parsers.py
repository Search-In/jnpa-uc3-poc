"""Gate Document UPLOAD parsers — EIR / PIN ticket / Form-13.

Pure functions that turn an uploaded CSV/XLS/XLSX byte payload into validated,
mapped records plus a preview and user-friendly errors — WITHOUT touching the DB.
Mirrors :mod:`services.cfs_ecy.upload_parsers` (alias-driven column mapping, the
same ParseResult envelope, the same byte readers) so the Data-Upload UX is
identical across modules.

Client-document reality this encodes:
  * EIR      — container, seal, vessel+VIA, BAT lane, TruckIn/TruckOut (-> TAT),
               To/From CFS, company, scanner stamp.
  * PIN      — PIN no, yard location in FREE format (e.g. '2P08D.1'), gate,
               company, move type; a dual-move ticket is two legs sharing the
               PIN number (leg_seq 1/2).
  * FORM-13  — VisitID, in/out gate codes (IGTK01/OGTK05), transporter, direction.

Container number is OPTIONAL for every doc type: real documents exist with no
container number (truck MH46AF4375). The TRUCK number is the required key
instead, so such documents are still ingested and stay truck-keyed.
"""
from __future__ import annotations

import csv
import datetime as _dt
import hashlib
import io
import re
from typing import Any, Optional

from jnpa_shared.iso6346 import is_valid_container_no

# JNPA operates in IST; document timestamps carry no timezone.
IST = _dt.timezone(_dt.timedelta(hours=5, minutes=30))

DOC_TYPES = ("EIR", "PIN", "FORM13")
MOVE_TYPES = ("IMPORT_PICK", "EXPORT_DROP", "EMPTY_PICK", "EMPTY_DROP")
DIRECTIONS = ("IMPORT", "EXPORT")


# ---------------------------------------------------------------- normalisers
def norm_header(name: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(name).strip().lower())


def clean(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    v = str(raw).strip()
    return v or None


def norm_container(raw: Any) -> Optional[str]:
    v = clean(raw)
    return v.upper().replace(" ", "") if v else None


def norm_plate(raw: Any) -> Optional[str]:
    v = clean(raw)
    return re.sub(r"[^A-Z0-9]", "", v.upper()) if v else None


def norm_move_type(raw: Any) -> Optional[str]:
    v = (clean(raw) or "").upper().replace("-", "_").replace(" ", "_")
    if not v:
        return None
    aliases = {
        "IMPORT_PICK": "IMPORT_PICK", "IMPORTPICKUP": "IMPORT_PICK",
        "PICK_UP_IMPORT": "IMPORT_PICK", "PICKUP_IMPORT": "IMPORT_PICK",
        "IMPORT": "IMPORT_PICK", "PICK": "IMPORT_PICK",
        "EXPORT_DROP": "EXPORT_DROP", "EXPORTDROP": "EXPORT_DROP",
        "DROP_EXPORT": "EXPORT_DROP", "EXPORT": "EXPORT_DROP", "DROP": "EXPORT_DROP",
        "EMPTY_PICK": "EMPTY_PICK", "EMPTYPICK": "EMPTY_PICK", "MTY_PICK": "EMPTY_PICK",
        "EMPTY_DROP": "EMPTY_DROP", "EMPTYDROP": "EMPTY_DROP", "MTY_DROP": "EMPTY_DROP",
    }
    return aliases.get(v.replace("__", "_"))


def norm_direction(raw: Any) -> Optional[str]:
    v = (clean(raw) or "").upper()
    if v in ("IMPORT", "IMP", "I", "INBOUND"):
        return "IMPORT"
    if v in ("EXPORT", "EXP", "E", "OUTBOUND"):
        return "EXPORT"
    return None


_TS_FORMATS = (
    "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%d-%m-%Y %H:%M:%S",
    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d",
    "%d-%b-%Y %H:%M", "%d-%b-%Y", "%d %b %Y %H:%M",
)


def parse_ts(raw: Any) -> Optional[_dt.datetime]:
    if raw is None:
        return None
    if isinstance(raw, _dt.datetime):
        return raw if raw.tzinfo else raw.replace(tzinfo=IST)
    if isinstance(raw, _dt.date):
        return _dt.datetime(raw.year, raw.month, raw.day, tzinfo=IST)
    s = clean(raw)
    if not s:
        return None
    for fmt in _TS_FORMATS:
        try:
            return _dt.datetime.strptime(s, fmt).replace(tzinfo=IST)
        except ValueError:
            continue
    try:
        d = _dt.datetime.fromisoformat(s.replace("Z", "+00:00"))
        return d if d.tzinfo else d.replace(tzinfo=IST)
    except ValueError:
        return None


def parse_num(raw: Any) -> Optional[float]:
    v = clean(raw)
    if v is None:
        return None
    try:
        return float(re.sub(r"[^0-9.\-]", "", v) or "0")
    except ValueError:
        return None


# ---------------------------------------------------------------- aliases
_COMMON = {
    "container_number": ("containerno", "containernumber", "container", "cntrno",
                         "cntr", "equipmentno", "boxno", "unitno"),
    "terminal": ("terminal", "terminalcode", "terminalname", "facility"),
    "remarks": ("remarks", "remark", "notes", "note", "stamps"),
}

ALIASES: dict[str, dict[str, tuple[str, ...]]] = {
    "EIR": {
        **_COMMON,
        "eir_no": ("eirno", "eirnumber", "eir", "documentno", "docno"),
        "eir_type": ("eirtype", "type", "movementtype", "transactiontype"),
        "vessel": ("vessel", "vesselname", "vessalname"),
        "via_no": ("via", "viano", "vianumber", "voyage", "voyageno"),
        "seal_number": ("seal", "sealno", "sealnumber", "conseal"),
        "bat_lane": ("bat", "batlane", "batno", "lane", "laneno", "batcode"),
        "truck_no": ("truckno", "trucknumber", "vehicleno", "vehiclenumber",
                     "lorryno", "trailerno", "plate", "plateno"),
        "driver_name": ("drivername", "driver"),
        "driver_licence": ("driverlicence", "driverlicense", "dl", "dlno",
                           "licenceno", "licenseno"),
        "truck_in_time": ("truckin", "truckintime", "gatein", "gateintime",
                          "intime", "arrivaltime"),
        "truck_out_time": ("truckout", "truckouttime", "gateout", "gateouttime",
                           "outtime", "departuretime"),
        "gross_weight_mt": ("grossweight", "grossweightmt", "weight", "weightmt", "grosswt"),
        "company": ("company", "companyname", "transporter", "transportername"),
        "cfs_from": ("from", "cfsfrom", "fromcfs", "origin"),
        "cfs_to": ("to", "cfsto", "tocfs", "destination"),
        "group_code": ("group", "groupcode", "grp"),
        "scanner_stamp": ("scannerstamp", "scanstamp", "scanner", "scanremark", "scanstatus"),
    },
    "PIN": {
        **_COMMON,
        "pin_number": ("pin", "pinno", "pinnumber", "pinticket", "ticketno", "ticketnumber"),
        "ticket_type": ("tickettype", "type", "visittype"),
        "truck_no": ("truckno", "trucknumber", "vehicleno", "vehiclenumber",
                     "lorryno", "trailerno", "plate", "plateno"),
        "company": ("company", "companyname", "transporter", "transportername"),
        "group_code": ("group", "groupcode", "grp", "emptygroup"),
        "yard_location": ("yard", "yardlocation", "yardloc", "location", "slot",
                          "position", "yardposition"),
        "gate": ("gate", "gateno", "gatenumber", "gatecode"),
        "move_type": ("movetype", "move", "moveedirection", "transactiontype", "movement"),
        "leg_seq": ("leg", "legseq", "legno", "sequence", "seq"),
        "issued_at": ("issuedat", "issuedon", "issuetime", "timestamp", "datetime",
                      "date", "tickettime"),
    },
    "FORM13": {
        **_COMMON,
        "form13_no": ("form13no", "form13", "formno", "documentno", "docno"),
        "visit_id": ("visitid", "visit", "visitno", "visitnumber"),
        "vehicle_no": ("vehicleno", "vehiclenumber", "truckno", "trucknumber",
                       "lorryno", "trailerno", "plate", "plateno"),
        "transporter_name": ("transporter", "transportername", "company", "companyname"),
        "driver_name": ("drivername", "driver"),
        "driver_licence": ("driverlicence", "driverlicense", "dl", "dlno",
                           "licenceno", "licenseno"),
        "in_gate": ("ingate", "ingatecode", "gatein", "ingateno"),
        "out_gate": ("outgate", "outgatecode", "gateout", "outgateno"),
        "direction": ("direction", "importexport", "impexp", "flow"),
        "bat_lane": ("bat", "batlane", "batno", "lane", "batcode"),
        "shipping_bill_no": ("shippingbillno", "shippingbill", "sbno", "sb"),
        "gross_wt_kg": ("grossweight", "grosswtkg", "weightkg", "grosswt", "weight"),
        "issued_at": ("issuedat", "issuedon", "issuetime", "timestamp", "datetime",
                      "date", "form13date"),
    },
}

# Required columns per doc type — the TRUCK, never the container (containerless
# documents are real and must remain ingestible).
_REQUIRED: dict[str, dict[str, tuple[str, ...]]] = {
    "EIR": {"Truck Number": ALIASES["EIR"]["truck_no"]},
    "PIN": {"PIN Number": ALIASES["PIN"]["pin_number"],
            "Truck Number": ALIASES["PIN"]["truck_no"]},
    "FORM13": {"Vehicle Number": ALIASES["FORM13"]["vehicle_no"]},
}

_TEMPLATES: dict[str, tuple[list[str], list[str], str]] = {
    "EIR": (
        ["EIR No", "Terminal", "Container No", "Vessel", "VIA", "Seal No", "BAT",
         "Truck No", "Driver Name", "Driver Licence", "Truck In", "Truck Out",
         "Company", "From CFS", "To CFS", "Scanner Stamp", "Remarks"],
        ["4339869", "PSA BMCT", "MSMU1908508", "SAV", "S0696", "EU31716082", "B723",
         "MH43BX1488", "BABALU KUMAR", "UP6420140008203", "06/06/2026 08:26",
         "06/06/2026 11:11", "TRANSTA", "CLP CFS", "", "SCANNED CLEAN", ""],
        "REQUIRED: Truck No. Container No is OPTIONAL (containerless EIRs are valid). "
        "Truck In/Out format DD/MM/YYYY HH:MM (IST) — TAT is computed from them.",
    ),
    "PIN": (
        ["PIN Number", "Terminal", "Truck No", "Company", "Container No", "Group Code",
         "Yard Location", "Gate", "Move Type", "Leg", "Issued At", "Remarks"],
        ["230283", "NSFT", "MH43CQ2814", "TRANSTAR", "OOLU9340457", "", "2P08D.1",
         "10", "IMPORT_PICK", "1", "10/06/2026 12:29", ""],
        "REQUIRED: PIN Number + Truck No. A DUAL-MOVE ticket is TWO rows sharing the "
        "PIN Number with Leg 1 and Leg 2. Move Type: IMPORT_PICK / EXPORT_DROP / "
        "EMPTY_PICK / EMPTY_DROP. Yard Location is free-format (e.g. 2P08D.1).",
    ),
    "FORM13": (
        ["Form13 No", "Visit ID", "Terminal", "Container No", "Vehicle No",
         "Transporter", "In Gate", "Out Gate", "Direction", "BAT", "Shipping Bill No",
         "Gross Wt (kg)", "Issued At", "Remarks"],
        ["F13000000001", "4418958", "NSIGT", "FFAU4770682", "MH43BX1488",
         "Transtar Handling & Warehousing Co", "IGTK01", "OGTK05", "IMPORT", "B723",
         "", "", "10/06/2026 12:29", ""],
        "REQUIRED: Vehicle No. Direction: IMPORT / EXPORT. In/Out Gate are the "
        "terminal gate codes (e.g. IGTK01 / OGTK05).",
    ),
}


def doc_type_ok(value: Optional[str]) -> Optional[str]:
    v = (value or "").strip().upper()
    return v if v in DOC_TYPES else None


def template_csv(doc_type: str) -> str:
    cols, example, guidance = _TEMPLATES[doc_type]
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(cols)
    w.writerow([f"# {guidance} Column names are flexible. Delete this line and the "
                "example row before uploading."] + [""] * (len(cols) - 1))
    w.writerow(example)
    return buf.getvalue()


# ---------------------------------------------------------------- byte readers
def read_rows_from_bytes(content: bytes, filename: str) -> tuple[list[str], list[dict[str, Any]]]:
    """Return (header, rows) from a CSV / XLS / XLSX payload. Identical semantics to
    services.cfs_ecy.upload_parsers.read_rows_from_bytes."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            it = ws.iter_rows(values_only=True)
            try:
                header = [str(c).strip() if c is not None else "" for c in next(it)]
            except StopIteration:
                raise ValueError("empty_file")
            rows = []
            for values in it:
                if not any(v not in (None, "") for v in values):
                    continue
                rows.append({header[i]: (values[i] if i < len(values) else None)
                             for i in range(len(header))})
        finally:
            wb.close()
        return header, rows
    if name.endswith(".xls"):
        import xlrd
        book = xlrd.open_workbook(file_contents=content)
        sh = book.sheet_by_index(0)
        if sh.nrows == 0:
            raise ValueError("empty_file")
        header = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]
        rows = []
        for r in range(1, sh.nrows):
            values = [sh.cell_value(r, c) for c in range(sh.ncols)]
            if not any(str(v).strip() for v in values):
                continue
            rows.append({header[i]: values[i] for i in range(len(header))})
        return header, rows
    if name.endswith((".csv", ".txt")) or name == "":
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        all_rows = [r for r in reader if any((c or "").strip() for c in r)]
        if not all_rows:
            raise ValueError("empty_file")
        header = [c.strip() for c in all_rows[0]]
        rows = []
        for r in all_rows[1:]:
            if r and str(r[0]).strip().startswith("#"):
                continue
            rows.append({header[i]: (r[i] if i < len(r) else None) for i in range(len(header))})
        return header, rows
    raise ValueError("unsupported_format")


# ---------------------------------------------------------------- ParseResult
class ParseResult:
    def __init__(self) -> None:
        self.errors: list[dict[str, Any]] = []
        self.warnings: list[dict[str, Any]] = []
        self.records: list[dict[str, Any]] = []
        self.preview: list[dict[str, Any]] = []
        self.row_count = 0
        self.invalid_count = 0
        self.duplicate_count = 0
        self.rejected = False

    def err(self, row: Optional[int], col: Optional[str], code: str, detail: str, raw: Any = None):
        self.errors.append({"row_number": row, "column_name": col, "error_code": code,
                            "error_detail": detail,
                            "raw_value": (None if raw is None else str(raw))})

    def warn(self, row: Optional[int], col: Optional[str], code: str, detail: str):
        self.warnings.append({"row_number": row, "column_name": col,
                              "error_code": code, "error_detail": detail})


def check_required_columns(res: ParseResult, header: list[str], doc_type: str) -> bool:
    hset = {norm_header(h) for h in header if norm_header(h)}
    missing = [label for label, aliases in _REQUIRED[doc_type].items()
               if not any(a in hset for a in aliases)]
    if missing:
        for label in missing:
            res.err(None, label, "missing_column",
                    f"{label} column not found. Please download the latest template.")
        res.rejected = True
        return False
    return True


def _pick(row_norm: dict[str, Any], doc_type: str, canonical: str) -> Optional[str]:
    for src in ALIASES[doc_type].get(canonical, ()):
        if src in row_norm:
            v = clean(row_norm[src])
            if v is not None:
                return v
    return None


def _sha(*parts: Any) -> str:
    return hashlib.sha256("|".join("" if p is None else str(p) for p in parts).encode()).hexdigest()


# ---------------------------------------------------------------- main parse
def parse(header: list[str], rows: list[dict[str, Any]], *, doc_type: str,
          source_file: Optional[str] = None) -> ParseResult:
    """Validate + map one uploaded gate-document file into canonical records."""
    res = ParseResult()
    res.row_count = len(rows)
    if doc_type not in DOC_TYPES:
        res.err(None, None, "invalid_doc_type", f"{doc_type} is not one of {DOC_TYPES}")
        res.rejected = True
        return res
    if not check_required_columns(res, header, doc_type):
        return res

    seen: set[str] = set()
    for i, raw in enumerate(rows, start=1):
        rn = {norm_header(k): v for k, v in raw.items() if norm_header(k)}
        rec = _parse_row(res, rn, i, doc_type, source_file)
        if rec is None:
            continue
        if rec["row_sha256"] in seen:
            res.duplicate_count += 1
            res.warn(i, None, "duplicate_in_file",
                     "an identical document row already appears earlier in this file (skipped)")
            continue
        seen.add(rec["row_sha256"])
        res.records.append(rec)

    res.preview = [_preview_row(r, doc_type) for r in res.records[:20]]
    return res


def _container_fields(res: ParseResult, rn: dict, i: int, doc_type: str) -> tuple[Optional[str], Optional[bool]]:
    """Container number is optional everywhere; when present it is ISO-6346 checked
    (invalid -> imported + flagged, never rejected)."""
    cn = norm_container(_pick(rn, doc_type, "container_number"))
    if not cn:
        res.warn(i, "Container No", "no_container_number",
                 "document has no container number (kept, truck-keyed)")
        return None, None
    iso_valid = bool(is_valid_container_no(cn))
    if not iso_valid:
        res.warn(i, "Container No", "container_iso6346_invalid",
                 f"{cn} fails the ISO-6346 check digit (imported, flagged)")
    return cn, iso_valid


def _parse_row(res: ParseResult, rn: dict, i: int, doc_type: str,
               source_file: Optional[str]) -> Optional[dict[str, Any]]:
    if doc_type == "EIR":
        truck = norm_plate(_pick(rn, doc_type, "truck_no"))
        if not truck:
            res.err(i, "Truck Number", "empty_required", "truck number is empty")
            res.invalid_count += 1
            return None
        cn, iso = _container_fields(res, rn, i, doc_type)
        t_in = parse_ts(_pick(rn, doc_type, "truck_in_time"))
        t_out = parse_ts(_pick(rn, doc_type, "truck_out_time"))
        if t_in and t_out and t_out < t_in:
            res.err(i, "Truck Out", "out_before_in",
                    "truck-out time is before truck-in time")
            res.invalid_count += 1
            return None
        rec = {
            "eir_no": _pick(rn, doc_type, "eir_no"),
            "eir_type": _pick(rn, doc_type, "eir_type"),
            "terminal": _pick(rn, doc_type, "terminal"),
            "container_number": cn, "iso_valid": iso,
            "vessel": _pick(rn, doc_type, "vessel"),
            "via_no": _pick(rn, doc_type, "via_no"),
            "seal_number": _pick(rn, doc_type, "seal_number"),
            "bat_lane": _pick(rn, doc_type, "bat_lane"),
            "truck_no": truck,
            "driver_name": _pick(rn, doc_type, "driver_name"),
            "driver_licence": _pick(rn, doc_type, "driver_licence"),
            "truck_in_time": t_in, "truck_out_time": t_out,
            "gross_weight_mt": parse_num(_pick(rn, doc_type, "gross_weight_mt")),
            "company": _pick(rn, doc_type, "company"),
            "cfs_from": _pick(rn, doc_type, "cfs_from"),
            "cfs_to": _pick(rn, doc_type, "cfs_to"),
            "group_code": _pick(rn, doc_type, "group_code"),
            "scanner_stamp": _pick(rn, doc_type, "scanner_stamp"),
            "remarks": _pick(rn, doc_type, "remarks"),
            "source_file": source_file,
        }
        rec["row_sha256"] = _sha(rec["eir_no"], truck, cn, t_in, t_out, rec["terminal"])
        return rec

    if doc_type == "PIN":
        pin = clean(_pick(rn, doc_type, "pin_number"))
        truck = norm_plate(_pick(rn, doc_type, "truck_no"))
        if not pin:
            res.err(i, "PIN Number", "empty_required", "PIN number is empty")
            res.invalid_count += 1
            return None
        if not truck:
            res.err(i, "Truck Number", "empty_required", "truck number is empty")
            res.invalid_count += 1
            return None
        cn, iso = _container_fields(res, rn, i, doc_type)
        raw_move = _pick(rn, doc_type, "move_type")
        move = norm_move_type(raw_move)
        if raw_move and move is None:
            res.warn(i, "Move Type", "unknown_move_type",
                     f"move type '{raw_move}' not recognised (stored as unset)")
        group = _pick(rn, doc_type, "group_code")
        if not cn and not group:
            res.warn(i, "Container No", "no_container_or_group",
                     "ticket carries neither container number nor group code")
        try:
            leg = int(float(clean(_pick(rn, doc_type, "leg_seq")) or 1))
        except (TypeError, ValueError):
            leg = 1
        rec = {
            "pin_number": pin,
            "ticket_type": _pick(rn, doc_type, "ticket_type"),
            "terminal": _pick(rn, doc_type, "terminal"),
            "truck_no": truck,
            "company": _pick(rn, doc_type, "company"),
            "container_number": cn, "iso_valid": iso,
            "group_code": group,
            "yard_location": _pick(rn, doc_type, "yard_location"),
            "gate": _pick(rn, doc_type, "gate"),
            "move_type": move,
            "leg_seq": max(1, leg),
            "issued_at": parse_ts(_pick(rn, doc_type, "issued_at")),
            "remarks": _pick(rn, doc_type, "remarks"),
            "source_file": source_file,
        }
        rec["row_sha256"] = _sha(pin, rec["leg_seq"], truck, cn, group, rec["terminal"])
        return rec

    # FORM13
    vehicle = norm_plate(_pick(rn, doc_type, "vehicle_no"))
    if not vehicle:
        res.err(i, "Vehicle Number", "empty_required", "vehicle number is empty")
        res.invalid_count += 1
        return None
    cn, iso = _container_fields(res, rn, i, doc_type)
    raw_dir = _pick(rn, doc_type, "direction")
    direction = norm_direction(raw_dir)
    if raw_dir and direction is None:
        res.warn(i, "Direction", "unknown_direction",
                 f"direction '{raw_dir}' not recognised (stored as unset)")
    rec = {
        "form13_no": _pick(rn, doc_type, "form13_no"),
        "visit_id": _pick(rn, doc_type, "visit_id"),
        "terminal": _pick(rn, doc_type, "terminal"),
        "container_number": cn, "iso_valid": iso,
        "vehicle_no": vehicle,
        "transporter_name": _pick(rn, doc_type, "transporter_name"),
        "driver_name": _pick(rn, doc_type, "driver_name"),
        "driver_licence": _pick(rn, doc_type, "driver_licence"),
        "in_gate": _pick(rn, doc_type, "in_gate"),
        "out_gate": _pick(rn, doc_type, "out_gate"),
        "direction": direction,
        "bat_lane": _pick(rn, doc_type, "bat_lane"),
        "shipping_bill_no": _pick(rn, doc_type, "shipping_bill_no"),
        "gross_wt_kg": parse_num(_pick(rn, doc_type, "gross_wt_kg")),
        "issued_at": parse_ts(_pick(rn, doc_type, "issued_at")),
        "remarks": _pick(rn, doc_type, "remarks"),
        "source_file": source_file,
    }
    rec["row_sha256"] = _sha(rec["form13_no"], rec["visit_id"], vehicle, cn, rec["terminal"])
    return rec


def _preview_row(r: dict[str, Any], doc_type: str) -> dict[str, Any]:
    def ts(v):
        return v.strftime("%d/%m/%Y %H:%M") if isinstance(v, _dt.datetime) else ""
    if doc_type == "EIR":
        tat = ""
        if r.get("truck_in_time") and r.get("truck_out_time"):
            tat = str(round((r["truck_out_time"] - r["truck_in_time"]).total_seconds() / 60))
        return {"EIR No": r.get("eir_no") or "", "Truck": r["truck_no"],
                "Container": r.get("container_number") or "(none)",
                "BAT": r.get("bat_lane") or "", "In": ts(r.get("truck_in_time")),
                "Out": ts(r.get("truck_out_time")), "TAT (min)": tat}
    if doc_type == "PIN":
        return {"PIN": r["pin_number"], "Leg": r["leg_seq"], "Truck": r["truck_no"],
                "Container": r.get("container_number") or (r.get("group_code") or "(none)"),
                "Yard": r.get("yard_location") or "", "Gate": r.get("gate") or "",
                "Move": r.get("move_type") or ""}
    return {"Form13": r.get("form13_no") or "", "Visit ID": r.get("visit_id") or "",
            "Vehicle": r["vehicle_no"], "Container": r.get("container_number") or "(none)",
            "In Gate": r.get("in_gate") or "", "Out Gate": r.get("out_gate") or "",
            "Direction": r.get("direction") or ""}
