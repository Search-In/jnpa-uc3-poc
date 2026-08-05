"""Pilot_card_data.xlsx parser → core.pilotage (+ core.pilot). Pure, no DB.

The pilot-card workbook has three sheets — INWARD / OUTWARD / SHIFTING — and the sheet
name IS the movement_type. Each data row is one pilotage movement. Emits:

  * ``_target='pilot'``    — one record per distinct pilot_id (upserted into core.pilot)
  * ``_target='pilotage'`` — one record per row (inserted into core.pilotage)

Column mapping is ALIAS-DRIVEN (headers normalised to lowercase-alnum), so the per-sheet
spelling variants (``first_line_assured`` vs ``firstLineAssured``, ``vacated_at`` vs
``berth_vacated_at``) all resolve. Everything not promoted to a canonical column is kept
verbatim in ``extras`` (jsonb), including the raw berth codes — nothing is dropped.

Each pilotage record carries a ``row_sha256`` content hash over its normalized canonical
fields, so re-importing an identical row collapses on the uq_pilotage_row unique index
(migration 0047). berth/call resolution is deferred to the repository (resolve-or-NULL).
"""
from __future__ import annotations

import datetime as _dt
import hashlib
import io
import json
from typing import Any, Optional

from ..upload_parsers import ParseResult

_MOVEMENTS = {"INWARD", "OUTWARD", "SHIFTING"}

# canonical pilotage field -> accepted NORMALISED header names (first present wins).
_ALIASES: dict[str, tuple[str, ...]] = {
    "via_no": ("vianumber",),
    "imo_no": ("imo",),
    "vessel_name": ("vesselname",),
    "pilot_code": ("pilotid",),
    "vessel_condition": ("vesselcondition",),
    "draft_fwd_m": ("forwarddraft",),
    "draft_aft_m": ("aftdraft",),
    "pilot_boarded_at": ("pilotboardedat",),
    "first_line_at": ("firstlineassured", "securedat"),
    "all_fast_at": ("allfastline",),
    "pilot_disembarked_at": ("pilotdisembarkedat",),
    "berth_vacated_at": ("berthvacatedat", "vacatedat", "unberthedat"),
    "anchor_down_at": ("anchordownat",),
    "anchor_up_at": ("anchorupat",),
    "submitted_at": ("submittedon",),
}
# Canonical timestamp fields (coerced to datetime).
_TS_FIELDS = ("pilot_boarded_at", "first_line_at", "all_fast_at", "pilot_disembarked_at",
              "berth_vacated_at", "anchor_down_at", "anchor_up_at", "submitted_at")
_NUM_FIELDS = ("draft_fwd_m", "draft_aft_m")
# Plausible vessel draft window (metres). Real drafts are < ~30 m; the DB column is
# numeric(5,2) (max 999.99). The source occasionally carries a garbage draft (e.g.
# forward_draft=5075.0 for CAPE SYROS in the OUTWARD sheet — a single bad cell, not a
# column shift), which would overflow the column. Such a value is stored as NULL and
# the raw kept in extras — never remapped, never clamped, never dropped-with-the-row.
_MAX_DRAFT_M = 99.99
# Headers consumed into canonical columns (excluded from extras). Berth headers are
# consumed separately (kept as raw codes in extras + resolved downstream).
_CONSUMED = {n for names in _ALIASES.values() for n in names} | {"berth", "fromberth", "toberth"}


def _norm(h: Any) -> str:
    return "".join(ch for ch in str(h or "").strip().lower() if ch.isalnum())


def _clean(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return v  # datetime / number passthrough


def _text(v: Any) -> Optional[str]:
    """Coerce an identity cell to clean text. openpyxl hands numbers back for
    numeric-looking cells (IMO 9974292 → 9974292.0), so a whole float is rendered
    without the trailing '.0'."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, (int, float)):
        return str(v)
    s = str(v).strip()
    return s or None


def _to_dt(v: Any) -> Optional[_dt.datetime]:
    if isinstance(v, _dt.datetime):
        return v
    if isinstance(v, _dt.date):
        return _dt.datetime(v.year, v.month, v.day)
    s = _clean(v)
    if not isinstance(s, str):
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M", "%d-%m-%Y %H:%M",
                "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return _dt.datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _to_num(v: Any) -> Optional[float]:
    s = _clean(v)
    if s is None:
        return None
    try:
        return float(str(s).replace(",", ""))
    except ValueError:
        return None


def _jsonable(v: Any) -> Any:
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.isoformat()
    return v


def _row_hash(rec: dict[str, Any]) -> str:
    keys = ("movement_type", "via_no", "imo_no", "vessel_name", "pilot_code",
            *_TS_FIELDS, *_NUM_FIELDS, "from_berth_code", "to_berth_code")
    payload = {k: _jsonable(rec.get(k)) for k in keys}
    return hashlib.sha256(json.dumps(payload, sort_keys=True, default=str).encode()).hexdigest()


def _berths(movement: str, cell: dict[str, Any]) -> tuple[Optional[str], Optional[str]]:
    """(from_berth_code, to_berth_code) by movement. INWARD arrives -> to; OUTWARD
    departs -> from; SHIFTING has both."""
    berth = _clean(cell.get("berth"))
    if movement == "INWARD":
        return None, berth
    if movement == "OUTWARD":
        return berth, None
    return _clean(cell.get("fromberth")), _clean(cell.get("toberth"))


def parse_pilot_card(content: bytes, filename: Optional[str] = None) -> ParseResult:
    res = ParseResult()
    try:
        import openpyxl  # noqa: PLC0415 — optional dep; fail as a typed error, never a crash
    except ImportError:
        res.rejected = True
        res.err(None, None, "xlsx_unsupported", "openpyxl is not installed; cannot read XLSX")
        return res

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        res.rejected = True
        res.err(None, None, "xlsx_read_error", f"could not read workbook: {exc}")
        return res

    pilots: dict[str, dict[str, Any]] = {}
    seen: set[str] = set()
    for sheet in wb.sheetnames:
        movement = sheet.strip().upper()
        if movement not in _MOVEMENTS:
            res.warn(None, sheet, "sheet_skipped", f"sheet '{sheet}' is not a pilot-card movement")
            continue
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        header = [_norm(h) for h in next(it, [])]
        col = {h: i for i, h in enumerate(header) if h}

        def cell(row: tuple, name: str) -> Any:
            i = col.get(name)
            return row[i] if (i is not None and i < len(row)) else None

        for rownum, row in enumerate(it, start=2):
            if not row:
                continue
            cells = {h: cell(row, h) for h in col}
            if _clean(cells.get("imo")) is None and _clean(cells.get("vianumber")) is None:
                continue  # blank trailing row
            res.row_count += 1

            def pick(field: str) -> Any:
                for src in _ALIASES.get(field, ()):
                    if src in cells and _clean(cells[src]) is not None:
                        return cells[src]
                return None

            from_code, to_code = _berths(movement, cells)
            rec: dict[str, Any] = {
                "_target": "pilotage", "_message": "PILOTAGE", "_source_file": filename,
                "movement_type": movement,
                "via_no": _text(pick("via_no")),
                "imo_no": _text(pick("imo_no")),
                "vessel_name": _text(pick("vessel_name")),
                "pilot_code": _text(pick("pilot_code")),
                "vessel_condition": _text(pick("vessel_condition")),
                "from_berth_code": _text(from_code),
                "to_berth_code": _text(to_code),
            }
            # Drafts: keep only plausible metres; an out-of-range value (source dirt)
            # is nulled to avoid a numeric(5,2) overflow, with the raw kept in extras.
            draft_rejects: dict[str, Any] = {}
            for f in _NUM_FIELDS:
                raw = _to_num(pick(f))
                if raw is None or 0 <= raw <= _MAX_DRAFT_M:
                    rec[f] = raw
                else:
                    rec[f] = None
                    draft_rejects[f] = raw
                    res.warn(rownum, f, "draft_out_of_range",
                             f"{f}={raw} outside 0..{_MAX_DRAFT_M} m; stored NULL (raw kept in extras)")
            for f in _TS_FIELDS:
                rec[f] = _to_dt(pick(f))
            # extras: every non-consumed, non-empty cell (verbatim) + raw berth codes
            # + any out-of-range draft that was nulled above.
            extras = {h: _jsonable(v) for h, v in cells.items()
                      if h not in _CONSUMED and _clean(v) is not None}
            if from_code:
                extras["raw_from_berth"] = from_code
            if to_code:
                extras["raw_to_berth"] = to_code
            for f, raw in draft_rejects.items():
                extras[f"raw_{f}"] = raw
            rec["extras"] = extras

            rec["row_sha256"] = _row_hash(rec)
            if rec["row_sha256"] in seen:
                res.duplicate_count += 1
                continue
            seen.add(rec["row_sha256"])
            res.records.append(rec)

            pc = rec["pilot_code"]
            if pc and pc not in pilots:
                pilots[pc] = {"_target": "pilot", "_message": "PILOTAGE",
                              "_source_file": filename, "pilot_code": pc, "name": None}

    # Emit pilots first (persist upserts them before pilotage, satisfying the FK).
    res.records = list(pilots.values()) + res.records
    res.preview = [{
        "Movement": r["movement_type"], "VIA": r.get("via_no") or "—",
        "Vessel": r.get("vessel_name") or "—", "IMO": r.get("imo_no") or "—",
        "Pilot": r.get("pilot_code") or "—",
        "Boarded": r["pilot_boarded_at"].strftime("%d/%m/%Y %H:%M") if r.get("pilot_boarded_at") else "—",
    } for r in res.records if r.get("_target") == "pilotage"][:20]
    return res
