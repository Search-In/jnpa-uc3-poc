"""Performance upload — PDF/CSV/XLSX parsing, validation, and mapping to perf_* rows.

Pure functions (no DB, no HTTP). Reads an uploaded file for one of the three report
types, validates columns/dates/numbers/terminals, and maps valid rows into
insert-ready records for the EXISTING jnpa.perf_* tables. Terminal codes are
normalised the same way as the PDF importer (GTI→APMT, BMCTPL→BMCT).

Two input shapes, ONE output contract (:class:`ParseResult`), so validate / preview /
import are identical whichever the client uploads:

* **PDF** — the official JNPA report exactly as published. Routed to
  :mod:`services.performance.pdf_parsers` (the validated extraction engine, shared
  with the offline backfill script). This is the client's primary workflow.
* **CSV / XLSX** — the normalised template produced by ``/templates/{report_type}``.

Format is decided by CONTENT (magic bytes) first and only then by extension, so a
mislabelled file is routed correctly instead of crashing the CSV reader.

For daily_status and monthly_teu the CSV/XLSX path also derives the JN_PORT aggregate
row (and the daily TOTAL status row) from the uploaded terminal rows, so an upload
immediately drives the dashboard KPI cards (which read the JN_PORT / TOTAL rollups).
The PDF path does NOT derive them — the official report prints its own JN Port /
TOTAL columns and those authoritative figures are extracted verbatim.
"""
from __future__ import annotations

import csv
import datetime as dt
import io
import re
from typing import Any, Optional

from . import pdf_parsers as PDF

IST = dt.timezone(dt.timedelta(hours=5, minutes=30))
MONTHS = {"JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
          "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12}

_TERM_ALIAS = {
    "NSFT": "NSFT", "NSICT": "NSICT", "NSIGT": "NSIGT", "NSDT": "NSDT", "JNPCT": "JNPCT",
    "GTI": "APMT", "APM": "APMT", "APMT": "APMT",
    "BMCT": "BMCT", "BMCTPL": "BMCT", "BMCTPSA": "BMCT", "PSA": "BMCT",
    "JN PORT": "JN_PORT", "JNPORT": "JN_PORT", "JN_PORT": "JN_PORT",
}
_CONTAINER_TERMINALS = ("NSFT", "NSICT", "NSIGT", "APMT", "BMCT", "NSDT")


def norm_terminal(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    s = str(raw).strip()
    m = re.search(r"\(([A-Za-z]+)\)", s)
    if m and m.group(1).upper() in _TERM_ALIAS:
        return _TERM_ALIAS[m.group(1).upper()]
    return _TERM_ALIAS.get(s.upper())


def _num(raw: Any) -> tuple[Optional[float], bool]:
    """(value, ok). Blank -> (None, True); non-numeric -> (None, False)."""
    if raw is None:
        return None, True
    s = str(raw).replace(",", "").strip()
    if s == "" or s in ("-", "–", "—"):
        return None, True
    try:
        return float(s), True
    except ValueError:
        return None, False


def _int(raw: Any) -> tuple[Optional[int], bool]:
    v, ok = _num(raw)
    return (int(round(v)) if v is not None else None), ok


def _date(raw: Any) -> Optional[dt.date]:
    if raw is None:
        return None
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# --- report-type specifications ---------------------------------------------
DAILY_STATUS_COLS = [
    "report_date", "terminal_code", "vessels", "imp_teus", "exp_teus", "total_teus",
    "yard_import_teus", "yard_export_teus", "yard_transhipment_teus", "yard_total_teus",
    "yard_usable_capacity_teus", "yard_occupancy_pct", "icd_pendency_teus",
    "cfs_pendency_teus", "gate_in_teus", "gate_out_teus", "gate_total_teus",
    "reefer_total_slots", "reefer_occupied_slots", "reefer_available_slots",
]
MONTHLY_TEU_COLS = ["month_date", "terminal_code", "vessel_calls", "discharge_teus",
                    "load_teus", "total_teus"]
LDB_REPORT_COLS = ["report_month", "terminal_code", "cycle", "segment",
                   "dwell_hours", "dwell_hours_prev"]

SPECS: dict[str, dict[str, Any]] = {
    "daily_status": {"columns": DAILY_STATUS_COLS, "required": ["report_date", "terminal_code"],
                     "example": ["2026-06-01", "NSFT", "2", "1775", "1748", "3523",
                                 "5859", "6729", "5720", "18308", "23433", "78.13",
                                 "1153", "1988", "1588", "1428", "3016", "1104", "548", "556"]},
    "monthly_teu": {"columns": MONTHLY_TEU_COLS, "required": ["month_date", "terminal_code"],
                    "example": ["2026-06-01", "NSFT", "48", "55951", "56326", "112277"]},
    "ldb_report": {"columns": LDB_REPORT_COLS,
                   "required": ["report_month", "terminal_code", "cycle", "segment"],
                   "example": ["2026-06-01", "NSFT", "IMPORT", "OVERALL", "22.8", "29.3"]},
}


def template_csv(report_type: str) -> str:
    spec = SPECS[report_type]
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(spec["columns"])
    w.writerow(spec["example"])
    return buf.getvalue()


# --- format detection --------------------------------------------------------
_PDF_MAGIC = b"%PDF"
_ZIP_MAGIC = b"PK\x03\x04"          # XLSX/XLSM are ZIP containers
_XLS_MAGIC = b"\xd0\xcf\x11\xe0"    # legacy OLE2 .xls


def sniff_format(content: bytes, filename: str) -> str:
    """Decide PDF / XLSX / CSV from the file's CONTENT, falling back to its extension.

    Magic bytes win because clients routinely rename reports (a PDF saved as .csv used
    to reach the CSV reader and blow up with a 500). Returns 'PDF' | 'XLSX' | 'CSV',
    or raises ValueError('unsupported_format') for anything we cannot read.
    """
    head = (content or b"")[:8]
    if head.startswith(_PDF_MAGIC):
        return "PDF"
    if head.startswith(_ZIP_MAGIC):
        return "XLSX"
    if head.startswith(_XLS_MAGIC):
        # Legacy .xls has no reader here; say so explicitly rather than mis-parsing it.
        raise ValueError("unsupported_format: legacy .xls is not supported — "
                         "save as .xlsx or .csv, or upload the original PDF")
    name = (filename or "").lower()
    if name.endswith(".pdf"):
        # Extension says PDF but the bytes do not — corrupt or truncated upload.
        raise ValueError("unreadable_pdf: file is named .pdf but does not start with "
                         "'%PDF' — it is corrupted or not a PDF")
    if name.endswith((".xlsx", ".xlsm")):
        raise ValueError("unreadable_file: file is named .xlsx but is not a valid "
                         "Excel workbook")
    if name.endswith((".csv", ".txt")) or name == "":
        return "CSV"
    raise ValueError("unsupported_format")


def is_pdf(content: bytes, filename: str = "") -> bool:
    return (content or b"").startswith(_PDF_MAGIC) or (filename or "").lower().endswith(".pdf")


# --- reading -----------------------------------------------------------------
def read_rows(content: bytes, filename: str) -> tuple[list[str], list[dict[str, Any]]]:
    """Return (header, rows) from a CSV or XLSX byte payload. Raises ValueError on
    an unreadable/empty file or unsupported format — never a bare csv/openpyxl error,
    so the upload layer answers 400/REJECTED instead of 500."""
    fmt = sniff_format(content, filename)
    if fmt == "PDF":
        # Callers must route PDFs to parse_pdf(); reaching here is a programming error.
        raise ValueError("unsupported_format: PDF must be parsed via parse_pdf()")
    if fmt == "XLSX":
        try:
            import openpyxl
        except ImportError as exc:
            raise ValueError("unreadable_file: XLSX support is unavailable on the "
                             "server (openpyxl is not installed)") from exc
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        except Exception as exc:  # noqa: BLE001 — corrupt/encrypted workbook
            raise ValueError(f"unreadable_file: {exc}") from exc
        try:
            ws = wb[wb.sheetnames[0]]
            it = ws.iter_rows(values_only=True)
            try:
                header = [str(c).strip() if c is not None else "" for c in next(it)]
            except StopIteration:
                raise ValueError("empty_file")
            rows = []
            for values in it:
                if not any(v not in (None, "") for v in values):
                    continue
                rows.append({header[i]: (values[i] if i < len(values) else None)
                             for i in range(len(header))})
        finally:
            wb.close()
        return header, rows
    # CSV
    text = content.decode("utf-8-sig", errors="replace").replace("\x00", "")
    try:
        reader = csv.reader(io.StringIO(text, newline=""))
        all_rows = [r for r in reader if any((c or "").strip() for c in r)]
    except csv.Error as exc:        # malformed CSV → clean rejection, not a 500
        raise ValueError(f"unreadable_file: malformed CSV ({exc})") from exc
    if not all_rows:
        raise ValueError("empty_file")
    header = [c.strip() for c in all_rows[0]]
    rows = []
    for r in all_rows[1:]:
        # skip commented/example helper lines
        if r and str(r[0]).strip().startswith("#"):
            continue
        rows.append({header[i]: (r[i] if i < len(r) else None) for i in range(len(header))})
    return header, rows


# --- validation + mapping ----------------------------------------------------
class ParseResult:
    def __init__(self) -> None:
        self.errors: list[dict[str, Any]] = []
        self.warnings: list[dict[str, Any]] = []
        self.records: dict[str, list[dict[str, Any]]] = {}
        self.preview: list[dict[str, Any]] = []
        self.report_keys: set = set()   # report_date / month_date / report_month values
        self.row_count = 0
        self.rejected = False           # structural failure (wrong template)

    def err(self, row: Optional[int], col: Optional[str], code: str, detail: str, raw: Any = None):
        self.errors.append({"row_number": row, "column_name": col, "error_code": code,
                            "error_detail": detail, "raw_value": (None if raw is None else str(raw))})

    def warn(self, row: Optional[int], col: Optional[str], code: str, detail: str):
        self.warnings.append({"row_number": row, "column_name": col, "error_code": code,
                              "error_detail": detail})


def _check_header(res: ParseResult, report_type: str, header: list[str]) -> bool:
    expected = SPECS[report_type]["columns"]
    hset, eset = set(header), set(expected)
    missing = [c for c in expected if c not in hset]
    if missing:
        res.err(None, None, "missing_columns",
                f"template mismatch — missing columns: {', '.join(missing)}")
        res.rejected = True
        return False
    return True


def parse(report_type: str, header: list[str], rows: list[dict[str, Any]]) -> ParseResult:
    res = ParseResult()
    if report_type not in SPECS:
        res.err(None, None, "unknown_report_type", f"unknown report type '{report_type}'")
        res.rejected = True
        return res
    if not _check_header(res, report_type, header):
        return res
    res.row_count = len(rows)
    if report_type == "daily_status":
        _parse_daily(res, rows)
    elif report_type == "monthly_teu":
        _parse_monthly(res, rows)
    else:
        _parse_ldb(res, rows)
    return res


def _numcols(res: ParseResult, rownum: int, row: dict, cols: list[str]) -> dict[str, Optional[float]]:
    out: dict[str, Optional[float]] = {}
    for c in cols:
        v, ok = _num(row.get(c))
        if not ok:
            res.err(rownum, c, "invalid_number", f"'{row.get(c)}' is not a number", row.get(c))
        out[c] = v
    return out


def _parse_daily(res: ParseResult, rows: list[dict]) -> None:
    traffic, status = [], []
    for i, row in enumerate(rows, start=2):     # row 1 = header
        rd = _date(row.get("report_date"))
        term = norm_terminal(row.get("terminal_code"))
        if rd is None:
            res.err(i, "report_date", "invalid_date", f"'{row.get('report_date')}' is not a valid date", row.get("report_date"))
        if term is None:
            res.err(i, "terminal_code", "unknown_terminal", f"'{row.get('terminal_code')}' is not a known terminal", row.get("terminal_code"))
        nums = _numcols(res, i, row, [c for c in DAILY_STATUS_COLS if c not in ("report_date", "terminal_code")])
        occ = nums.get("yard_occupancy_pct")
        if occ is not None and (occ < 0 or occ > 100):
            res.warn(i, "yard_occupancy_pct", "occupancy_out_of_range", f"{occ}% outside 0–100")
        if rd is None or term is None:
            continue
        res.report_keys.add(rd)
        traffic.append({"report_date": rd, "terminal_code": term, "period": "DAY",
                        "vessels": (int(nums["vessels"]) if nums.get("vessels") is not None else None),
                        "imp_teus": nums.get("imp_teus"), "exp_teus": nums.get("exp_teus"),
                        "total_teus": nums.get("total_teus")})
        status.append({"report_date": rd, "terminal_code": term,
                       "icd_pendency_teus": nums.get("icd_pendency_teus"),
                       "cfs_pendency_teus": nums.get("cfs_pendency_teus"),
                       "yard_import_teus": nums.get("yard_import_teus"),
                       "yard_export_teus": nums.get("yard_export_teus"),
                       "yard_transhipment_teus": nums.get("yard_transhipment_teus"),
                       "yard_total_teus": nums.get("yard_total_teus"),
                       "yard_usable_capacity_teus": nums.get("yard_usable_capacity_teus"),
                       "yard_occupancy_pct": occ,
                       "gate_in_teus": nums.get("gate_in_teus"), "gate_out_teus": nums.get("gate_out_teus"),
                       "gate_total_teus": nums.get("gate_total_teus"),
                       "reefer_total_slots": (int(nums["reefer_total_slots"]) if nums.get("reefer_total_slots") is not None else None),
                       "reefer_occupied_slots": (int(nums["reefer_occupied_slots"]) if nums.get("reefer_occupied_slots") is not None else None),
                       "reefer_available_slots": (int(nums["reefer_available_slots"]) if nums.get("reefer_available_slots") is not None else None)})
    # derive JN_PORT traffic + TOTAL status rollups per report_date from real rows
    _daily_rollups(traffic, status)
    res.records = {"snapshot": [{"report_date": d} for d in sorted(res.report_keys)],
                   "traffic": traffic, "status": status}
    res.preview = [dict(r) for r in status[:20]]


def _sum(vals):
    xs = [v for v in vals if v is not None]
    return sum(xs) if xs else None


def _daily_rollups(traffic: list[dict], status: list[dict]) -> None:
    from collections import defaultdict
    by_date_t = defaultdict(list)
    for r in traffic:
        if r["terminal_code"] in _CONTAINER_TERMINALS:
            by_date_t[r["report_date"]].append(r)
    for d, rs in by_date_t.items():
        traffic.append({"report_date": d, "terminal_code": "JN_PORT", "period": "DAY",
                        "vessels": _sum(r.get("vessels") for r in rs),
                        "imp_teus": _sum(r.get("imp_teus") for r in rs),
                        "exp_teus": _sum(r.get("exp_teus") for r in rs),
                        "total_teus": _sum(r.get("total_teus") for r in rs)})
    by_date_s = defaultdict(list)
    for r in status:
        if r["terminal_code"] in _CONTAINER_TERMINALS:
            by_date_s[r["report_date"]].append(r)
    for d, rs in by_date_s.items():
        yt = _sum(r.get("yard_total_teus") for r in rs)
        uc = _sum(r.get("yard_usable_capacity_teus") for r in rs)
        status.append({"report_date": d, "terminal_code": "TOTAL",
                       "icd_pendency_teus": _sum(r.get("icd_pendency_teus") for r in rs),
                       "cfs_pendency_teus": _sum(r.get("cfs_pendency_teus") for r in rs),
                       "yard_import_teus": _sum(r.get("yard_import_teus") for r in rs),
                       "yard_export_teus": _sum(r.get("yard_export_teus") for r in rs),
                       "yard_transhipment_teus": _sum(r.get("yard_transhipment_teus") for r in rs),
                       "yard_total_teus": yt,
                       "yard_usable_capacity_teus": uc,
                       "yard_occupancy_pct": (round(yt / uc * 100, 2) if yt is not None and uc else None),
                       "gate_in_teus": _sum(r.get("gate_in_teus") for r in rs),
                       "gate_out_teus": _sum(r.get("gate_out_teus") for r in rs),
                       "gate_total_teus": _sum(r.get("gate_total_teus") for r in rs),
                       "reefer_total_slots": _sum(r.get("reefer_total_slots") for r in rs),
                       "reefer_occupied_slots": _sum(r.get("reefer_occupied_slots") for r in rs),
                       "reefer_available_slots": _sum(r.get("reefer_available_slots") for r in rs)})


def _parse_monthly(res: ParseResult, rows: list[dict]) -> None:
    monthly = []
    for i, row in enumerate(rows, start=2):
        md = _date(row.get("month_date"))
        term = norm_terminal(row.get("terminal_code"))
        if md is None:
            res.err(i, "month_date", "invalid_date", f"'{row.get('month_date')}' is not a valid date", row.get("month_date"))
        if term is None:
            res.err(i, "terminal_code", "unknown_terminal", f"'{row.get('terminal_code')}' is not a known terminal", row.get("terminal_code"))
        nums = _numcols(res, i, row, ["vessel_calls", "discharge_teus", "load_teus", "total_teus"])
        if md is None or term is None:
            continue
        md = md.replace(day=1)
        fy_start = md.year if md.month >= 4 else md.year - 1
        res.report_keys.add(md)
        monthly.append({"fiscal_year": f"FY-{fy_start}-{str(fy_start + 1)[2:]}",
                        "month_date": md, "year_label": str(md.year),
                        "month_label": [k for k, v in MONTHS.items() if v == md.month][0],
                        "terminal_code": term,
                        "vessel_calls": (int(nums["vessel_calls"]) if nums.get("vessel_calls") is not None else None),
                        "discharge_teus": nums.get("discharge_teus"),
                        "load_teus": nums.get("load_teus"), "total_teus": nums.get("total_teus")})
    # JN_PORT monthly rollup per month from real container-terminal rows
    from collections import defaultdict
    by_m = defaultdict(list)
    for r in monthly:
        if r["terminal_code"] in _CONTAINER_TERMINALS:
            by_m[r["month_date"]].append(r)
    for md, rs in by_m.items():
        fy_start = md.year if md.month >= 4 else md.year - 1
        monthly.append({"fiscal_year": f"FY-{fy_start}-{str(fy_start + 1)[2:]}", "month_date": md,
                        "year_label": str(md.year),
                        "month_label": [k for k, v in MONTHS.items() if v == md.month][0],
                        "terminal_code": "JN_PORT",
                        "vessel_calls": _sum(r.get("vessel_calls") for r in rs),
                        "discharge_teus": _sum(r.get("discharge_teus") for r in rs),
                        "load_teus": _sum(r.get("load_teus") for r in rs),
                        "total_teus": _sum(r.get("total_teus") for r in rs)})
    res.records = {"monthly": monthly}
    res.preview = [dict(r) for r in monthly[:20]]


def _parse_ldb(res: ParseResult, rows: list[dict]) -> None:
    ldb = []
    for i, row in enumerate(rows, start=2):
        rm = _date(row.get("report_month"))
        term = norm_terminal(row.get("terminal_code"))
        cycle = str(row.get("cycle") or "").strip().upper()
        segment = str(row.get("segment") or "").strip().upper()
        if rm is None:
            res.err(i, "report_month", "invalid_date", f"'{row.get('report_month')}' is not a valid date", row.get("report_month"))
        if term is None:
            res.err(i, "terminal_code", "unknown_terminal", f"'{row.get('terminal_code')}' is not a known terminal", row.get("terminal_code"))
        if cycle not in ("IMPORT", "EXPORT"):
            res.err(i, "cycle", "invalid_cycle", f"'{row.get('cycle')}' must be IMPORT or EXPORT", row.get("cycle"))
        if segment not in ("OVERALL", "TRUCK", "TRAIN"):
            res.err(i, "segment", "invalid_segment", f"'{row.get('segment')}' must be OVERALL/TRUCK/TRAIN", row.get("segment"))
        nums = _numcols(res, i, row, ["dwell_hours", "dwell_hours_prev"])
        if rm is None or term is None or cycle not in ("IMPORT", "EXPORT") or segment not in ("OVERALL", "TRUCK", "TRAIN"):
            continue
        rm = rm.replace(day=1)
        res.report_keys.add(rm)
        ldb.append({"report_month": rm, "terminal_code": term, "cycle": cycle, "segment": segment,
                    "dwell_hours": nums.get("dwell_hours"), "dwell_hours_prev": nums.get("dwell_hours_prev")})
    res.records = {"ldb_port_dwell": ldb}
    res.preview = [dict(r) for r in ldb[:20]]


# =============================================================================
# PDF path — the official JNPA reports, as published
# =============================================================================
# Content fingerprints identifying which official report a PDF actually is. Used to
# reject a right-file/wrong-report-type upload before it writes anything.
_PDF_SIGNATURES: dict[str, tuple[str, ...]] = {
    "daily_status": ("DAILY STATUS REPORT", "CONTAINER TERMINALS TRAFFIC"),
    "monthly_teu": ("MONTHWISE TEUS HANDLED", "CONTAINER TERMINALS"),
    "ldb_report": ("LOGISTICS DATA BANK", "NICDC LOGISTICS DATA SERVICES", "LDB"),
}
_REPORT_LABEL = {"daily_status": "Daily Status Report",
                 "monthly_teu": "FY JN Port TEUs Report",
                 "ldb_report": "NLDS/LDB Analytics Report"}

# Report-date recovery from the PDF BODY, used only when the filename carries no date
# (clients rename downloads). The parsers themselves read the date from the filename,
# so we hand them a synthetic canonical name rather than touching extraction logic.
_BODY_DATE = re.compile(r"As\s+on\s+(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})", re.I)
_BODY_DATE2 = re.compile(r"Date\s*:?\s*(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})", re.I)
_BODY_MON_YY = re.compile(r"\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*'?\s?(\d{2,4})\b")
_LDB_MONTH_NAMES = {1: "January", 2: "February", 3: "March", 4: "April", 5: "May", 6: "June",
                    7: "July", 8: "August", 9: "September", 10: "October", 11: "November",
                    12: "December"}


def pdf_text(content: bytes, *, max_pages: int = 6) -> str:
    """First-N-pages text of an uploaded PDF, for type/date detection.

    Raises ValueError('scanned_pdf') when the document has no text layer at all —
    those need OCR, which this pipeline does not perform.
    """
    with PDF.open_pdf(content) as pdf:
        pages = pdf.pages[:max_pages]
        text = "\n".join((p.extract_text() or "") for p in pages)
        if not text.strip() and not PDF._has_text(pdf):
            raise ValueError("scanned_pdf: this PDF has no text layer (it looks like a "
                             "scan or image export). Upload the original digital PDF "
                             "published by JNPA, or convert it with OCR first")
    if not text.strip():
        raise ValueError("empty_pdf: no readable text on the first pages")
    return text


def detect_pdf_report_type(text: str) -> Optional[str]:
    """Which official report this PDF is, from its content. None when inconclusive."""
    u = " ".join((text or "").upper().split())
    if "DAILY STATUS REPORT" in u:
        return "daily_status"
    if "MONTHWISE TEUS HANDLED" in u:
        return "monthly_teu"
    if "LOGISTICS DATA BANK" in u or "NICDC LOGISTICS DATA SERVICES" in u:
        return "ldb_report"
    return None


def _canonical_daily_name(filename: str, text: str) -> str:
    """A filename the validated parser can read a date from: the original when it
    already carries one, else one rebuilt from the report body."""
    if PDF.daily_date_from_name(filename or ""):
        return filename
    m = _BODY_DATE.search(text) or _BODY_DATE2.search(text)
    if not m:
        raise ValueError("missing_report_date: could not determine the report date from "
                         "the filename or the PDF body — rename the file to include it "
                         "(e.g. 'Daily Status Report 26.05.2026.pdf')")
    d, mth, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        dt.date(y, mth, d)
    except ValueError as exc:
        raise ValueError(f"invalid_report_date: '{m.group(0)}' is not a valid date") from exc
    return f"Daily Status Report {d:02d}.{mth:02d}.{y:04d}.pdf"


def _canonical_ldb_name(filename: str, text: str) -> str:
    """Same idea for the LDB report, whose parser expects '<Month>_<YYYY>'."""
    if PDF.ldb_month_from_name(filename or ""):
        return filename
    m = _BODY_MON_YY.search(text or "")
    if not m:
        raise ValueError("missing_report_month: could not determine the report month from "
                         "the filename or the PDF body — rename the file to include it "
                         "(e.g. 'NLDS_LDB_Full Analysis_March_2026.pdf')")
    mon = MONTHS[m.group(1).upper()]
    yr = int(m.group(2))
    yr += 2000 if yr < 100 else 0
    return f"LDB_{_LDB_MONTH_NAMES[mon]}_{yr}.pdf"


def parse_pdf(report_type: str, content: bytes, filename: str) -> ParseResult:
    """Parse an official JNPA report PDF into the SAME ParseResult the CSV/XLSX path
    produces, so validate / preview / import are unchanged.

    Extraction is delegated verbatim to :mod:`services.performance.pdf_parsers`; this
    function only detects the report, recovers the report date when the filename lacks
    it, and maps the parser output onto the repository's record keys.
    """
    res = ParseResult()
    if report_type not in SPECS:
        res.err(None, None, "unknown_report_type", f"unknown report type '{report_type}'")
        res.rejected = True
        return res

    text = pdf_text(content)                       # raises on scanned / unreadable
    detected = detect_pdf_report_type(text)
    if detected and detected != report_type:
        res.err(None, None, "report_type_mismatch",
                f"this PDF is a {_REPORT_LABEL[detected]}, but '{_REPORT_LABEL[report_type]}' "
                f"was selected — pick the matching report type and retry")
        res.rejected = True
        return res
    if detected is None:
        res.warn(None, None, "unrecognised_layout",
                 "could not fingerprint this PDF as an official JNPA report — parsing it "
                 f"as a {_REPORT_LABEL[report_type]}; check the preview carefully")

    if report_type == "daily_status":
        _pdf_daily(res, content, _canonical_daily_name(filename, text))
    elif report_type == "monthly_teu":
        _pdf_monthly(res, content, filename)
    else:
        _pdf_ldb(res, content, _canonical_ldb_name(filename, text))

    if not res.rejected and not any(res.records.values()):
        res.err(None, None, "no_data_extracted",
                "the PDF was read but no data rows could be extracted — it may be an "
                "unexpected layout or a different report")
        res.rejected = True
    return res


def _pdf_daily(res: ParseResult, content: bytes, filename: str) -> None:
    parsed = PDF.parse_daily(content, filename)
    if parsed is None:
        res.err(None, None, "missing_report_date",
                "could not determine the report date for this Daily Status Report")
        res.rejected = True
        return
    rd = parsed["report_date"]
    res.report_keys.add(rd)
    res.records = {
        "snapshot": [{"report_date": rd, "as_of_ts": parsed.get("as_of_ts"),
                      "source_file": parsed.get("source_file")}],
        "traffic": parsed["traffic"],
        "tonnage": parsed["tonnage"],
        "status": parsed["status"],
        "vessels": parsed["vessels"],
    }
    res.row_count = sum(len(v) for v in res.records.values())
    # Preview mirrors the report's own section order so the operator recognises it.
    res.preview = [dict(r) for r in parsed["status"][:20]]


def _pdf_monthly(res: ParseResult, content: bytes, filename: str) -> None:
    monthly = PDF.parse_monthly(content, filename)
    for r in monthly:
        res.report_keys.add(r["month_date"])
    res.records = {"monthly": monthly}
    res.row_count = len(monthly)
    res.preview = [dict(r) for r in monthly[:20]]


def _pdf_ldb(res: ParseResult, content: bytes, filename: str) -> None:
    ldb = PDF.parse_ldb(content, filename)
    if not ldb:
        res.err(None, None, "missing_report_month",
                "could not determine the report month for this LDB Analytics Report")
        res.rejected = True
        return
    res.records = {
        "ldb_port_dwell": ldb.get("port_dwell", []),
        "ldb_facility": ldb.get("facility", []),
        "ldb_congestion": ldb.get("congestion", []),
        "ldb_routes": ldb.get("routes", []),
        "ldb_weather": ldb.get("weather", []),
    }
    for r in res.records["ldb_port_dwell"]:
        res.report_keys.add(r["report_month"])
    res.row_count = sum(len(v) for v in res.records.values())
    res.preview = [dict(r) for r in res.records["ldb_port_dwell"][:20]]
