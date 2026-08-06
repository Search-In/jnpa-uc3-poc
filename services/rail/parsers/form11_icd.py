"""Form 11 (XLSX) + CTO (TXT) parser (group ``rail-form11-icd``).

This group carries three physical shapes; the parser dispatches on
extension + content:

  * **Form 11 XLSX** — an export pre-advice manifest, one row per container.
    Two header variants exist in the corpus (a terse `Container No / Iso Code
    / VGM Wt ...` sheet and the verbose `LINER_BOOKING_NUMBER / CONTAINER_NO
    ...` sheet); both collapse to one alias-mapped row. The **terminal**
    (BMCT / NSICT / NSIGT) is recovered FROM THE FILENAME.
  * **CTO TXT** — a headerless, comma-separated rail manifest, one wagon line.
    Two positional layouts exist (date-first vs rake-first); both are handled.
    The **cto_code** (e.g. R261076) is recovered FROM THE FILENAME.
  * **ICD daily-report PDF** — OUT OF SCOPE. Recognised and returned as
    ``feed='UNSUPPORTED'`` so the service rejects-not-crashes.

A stray FOIS Train Intimation CSV that lands in this group is detected by its
header and delegated to :mod:`services.rail.parsers.fois_csv`.
"""
from __future__ import annotations

import csv
import io
import re
from typing import Any, Optional

from jnpa_shared.iso6346 import is_valid_container_no

from . import (
    ParseResult, clean, norm_header, parse_date_time, to_float, to_int,
    upper_or_none,
)
from . import fois_csv

# ---------------------------------------------------------------- Form 11
# canonical field -> accepted NORMALISED header aliases.
FORM11_ALIASES: dict[str, tuple[str, ...]] = {
    "container_no": ("containerno", "containernumber", "cntrno"),
    "iso_code": ("isocode", "iso"),
    "box_size": ("size", "containersize"),
    "booking_number": ("linerbookingnumber", "bookingnumber", "booking"),
    "gross_weight": ("vgmwt", "grossweight", "vgm", "weight"),
    "pod": ("pod", "portofdischarge"),
    "line_code": ("linecode", "boxoperatorcode", "line"),
    "icd_location": ("icdlocation", "location"),
    "via": ("via",),
    "status": ("status", "origintype"),
}
_FORM11_KNOWN = {a for aliases in FORM11_ALIASES.values() for a in aliases}

# Terminal tokens the Form 11 filenames encode (BMCT FORM 11 / BT NSICT ...).
_TERMINALS = ("BMCT", "NSICT", "NSIGT", "NSFT", "GTIL", "NSDT")

_DATE_RE = re.compile(r"^\d{1,2}[-./]\d{1,2}[-./]\d{2,4}$")


def _terminal_from_filename(filename: str) -> str:
    upper = (filename or "").upper()
    for term in _TERMINALS:
        if term in upper:
            return term
    return "UNKNOWN"


def _cto_code_from_filename(filename: str) -> str:
    m = re.search(r"R\d{4,}", (filename or "").upper())
    if m:
        return m.group(0)
    stem = re.split(r"[\\/]", filename or "")[-1]
    stem = re.sub(r"\.[^.]+$", "", stem).strip()
    return (stem.split()[0].upper() if stem else "CTO")


def _cto_party_from_filename(filename: str) -> Optional[str]:
    stem = re.split(r"[\\/]", filename or "")[-1]
    stem = re.sub(r"\.[^.]+$", "", stem)
    # e.g. "R261076 HTPL" -> party token "HTPL".
    parts = [p for p in re.split(r"[\s_]+", stem) if p]
    for p in parts:
        if not re.match(r"^R\d{4,}$", p.upper()) and p.isalpha():
            return p.upper()
    return None


# ---------------------------------------------------------------- Form 11 read
def _read_xlsx(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True,
                                read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        try:
            header = [str(c).strip() if c is not None else ""
                      for c in next(it)]
        except StopIteration:
            return [], []
        rows = []
        for values in it:
            if not any(v not in (None, "") for v in values):
                continue
            rows.append({header[i]: (values[i] if i < len(values) else None)
                         for i in range(len(header))})
    finally:
        wb.close()
    return header, rows


def _pick(row_norm: dict[str, Any], canonical: str):
    for alias in FORM11_ALIASES.get(canonical, ()):
        if alias in row_norm:
            v = clean(row_norm[alias])
            if v is not None:
                return v
    return None


def _parse_form11(content: bytes, filename: str) -> ParseResult:
    res = ParseResult(feed="FORM11")
    terminal = _terminal_from_filename(filename)
    try:
        header, rows = _read_xlsx(content)
    except Exception as exc:  # noqa: BLE001 — corrupt workbook → rejection
        res.rejected = True
        res.reason = "unreadable_xlsx"
        res.err(None, None, "unreadable_file", f"could not read workbook: {exc}")
        return res

    res.row_count = len(rows)
    hset = {norm_header(h) for h in header if norm_header(h)}
    if not any(a in hset for a in FORM11_ALIASES["container_no"]):
        res.rejected = True
        res.reason = "missing_container_column"
        res.err(None, "Container No", "missing_column",
                "Container No column not found — not a Form 11 manifest.")
        return res

    for i, raw in enumerate(rows, start=1):
        row_norm = {norm_header(k): v for k, v in raw.items() if norm_header(k)}
        container = upper_or_none(_pick(row_norm, "container_no"))
        if not container:
            res.err(i, "Container No", "empty_container",
                    "container number is empty")
            res.invalid_count += 1
            continue
        iso_valid = bool(is_valid_container_no(container))
        if not iso_valid:
            res.warn(i, "Container No", "container_iso6346_invalid",
                     f"{container} fails the ISO-6346 check digit (imported, "
                     "flagged)")
        extra = {k: clean(v) for k, v in raw.items()
                 if norm_header(k) and norm_header(k) not in _FORM11_KNOWN
                 and clean(v) is not None}
        res.rows.append({
            "terminal": terminal,
            "container_no": container,
            "iso_code": clean(_pick(row_norm, "iso_code")),
            "box_size": clean(_pick(row_norm, "box_size")),
            "booking_number": clean(_pick(row_norm, "booking_number")),
            "gross_weight": to_float(_pick(row_norm, "gross_weight")),
            "pod": upper_or_none(_pick(row_norm, "pod")),
            "line_code": upper_or_none(_pick(row_norm, "line_code")),
            "icd_location": clean(_pick(row_norm, "icd_location")),
            "via": clean(_pick(row_norm, "via")),
            "status": clean(_pick(row_norm, "status")),
            "iso_valid": iso_valid,
            "source_file": filename,
            "extra": extra,
        })

    res.preview = [{"Terminal": r["terminal"], "Container": r["container_no"],
                    "ISO": r.get("iso_code"), "POD": r.get("pod"),
                    "Line": r.get("line_code")} for r in res.rows[:20]]
    return res


# ---------------------------------------------------------------- CTO read
_CONTAINER_RE = re.compile(r"^[A-Z]{4}\d{6,7}$")


def _read_lines(content: bytes) -> list[list[str]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return [r for r in reader if any((c or "").strip() for c in r)]


def _at(row: list[str], idx: int) -> Optional[str]:
    return clean(row[idx]) if idx < len(row) else None


def _norm_container(raw: Optional[str]) -> tuple[Optional[str], bool]:
    """(container_no, is_empty). 'EMPTY WAGON' / 'EMPTY' → (None, True)."""
    v = upper_or_none(raw)
    if not v or v.replace(" ", "") in ("EMPTYWAGON", "EMPTY", "NA"):
        return None, True
    return v.replace(" ", ""), False


def _parse_cto(rows: list[list[str]], filename: str) -> ParseResult:
    res = ParseResult(feed="CTO")
    res.row_count = len(rows)
    cto_code = _cto_code_from_filename(filename)
    party = _cto_party_from_filename(filename)

    for i, row in enumerate(rows, start=1):
        # Two positional layouts: date-first (A) vs rake-first (B).
        if _at(row, 1) and _DATE_RE.match(_at(row, 1) or ""):
            idx = {"seq": 0, "date": 1, "time": 2, "rake_no": 3, "wagon": 4,
                   "container": 5, "size": 6, "le": 7, "line": 8, "weight": 9,
                   "pol": 10, "pod": 11, "from": 12, "terminal": 13,
                   "ref": 14}
            rake_id = None
        else:
            idx = {"seq": 0, "rake_no": 1, "date": 2, "time": 3, "rake_id": 4,
                   "wagon": 5, "container": 6, "size": 7, "le": 8, "line": 9,
                   "weight": 10, "pol": 11, "pod": 12, "from": 13,
                   "terminal": 14, "ref": None}
            rake_id = _at(row, idx["rake_id"])

        wagon = upper_or_none(_at(row, idx["wagon"]))
        if not wagon:
            res.err(i, "wagon_no", "empty_wagon", "wagon number is empty",
                    ",".join(row))
            res.invalid_count += 1
            continue

        container, is_empty = _norm_container(_at(row, idx["container"]))
        iso_valid = bool(container and is_valid_container_no(container))
        if container and not iso_valid:
            res.warn(i, "container_no", "container_iso6346_invalid",
                     f"{container} fails the ISO-6346 check digit "
                     "(imported, flagged)")
        extra: dict[str, Any] = {"seq_raw": _at(row, idx["seq"])}
        if party:
            extra["party"] = party
        ref = _at(row, idx["ref"]) if idx.get("ref") is not None else None
        if ref:
            extra["booking_ref"] = ref

        res.rows.append({
            "cto_code": cto_code,
            "rake_no": upper_or_none(_at(row, idx["rake_no"])),
            "rake_id": upper_or_none(rake_id),
            "seq": to_int(_at(row, idx["seq"])),
            "wagon_no": wagon,
            "container_no": container,
            "is_empty": is_empty,
            "box_size": _at(row, idx["size"]),
            "load_empty": upper_or_none(_at(row, idx["le"])),
            "line_code": upper_or_none(_at(row, idx["line"])),
            "weight": to_float(_at(row, idx["weight"])),
            "pol": upper_or_none(_at(row, idx["pol"])),
            "pod": upper_or_none(_at(row, idx["pod"])),
            "from_station": upper_or_none(_at(row, idx["from"])),
            "terminal": upper_or_none(_at(row, idx["terminal"])),
            "booking_ref": ref,
            "event_ts": parse_date_time(_at(row, idx["date"]),
                                        _at(row, idx["time"])),
            "iso_valid": iso_valid if container else None,
            "source_file": filename,
            "extra": extra,
        })

    res.preview = [{"CTO": r["cto_code"], "Wagon": r["wagon_no"],
                    "Container": r["container_no"], "Size": r.get("box_size"),
                    "L/E": r.get("load_empty")} for r in res.rows[:20]]
    return res


# ---------------------------------------------------------------- dispatch
def _looks_like_fois(rows: list[list[str]]) -> bool:
    if not rows:
        return False
    header = {norm_header(c) for c in rows[0]}
    return any(a in header for a in fois_csv.ALIASES["rake_id"])


def parse(content: bytes, filename: str) -> ParseResult:
    name = (filename or "").lower()

    if name.endswith(".pdf"):
        res = ParseResult(feed="UNSUPPORTED")
        res.unsupported = True
        res.rejected = True
        res.reason = "UNSUPPORTED_FORMAT"
        res.err(None, None, "unsupported_format",
                "ICD daily-report PDF is not parsed (out of scope).")
        return res

    if name.endswith((".xlsx", ".xlsm", ".xls")):
        return _parse_form11(content, filename)

    if name.endswith((".txt", ".csv")):
        rows = _read_lines(content)
        if _looks_like_fois(rows):
            # a FOIS intimation that landed in this group — delegate.
            return fois_csv.parse(content, filename)
        if not rows:
            res = ParseResult(feed="CTO")
            res.reason = "empty_file"
            return res
        return _parse_cto(rows, filename)

    res = ParseResult(feed="UNSUPPORTED")
    res.unsupported = True
    res.rejected = True
    res.reason = "UNSUPPORTED_FORMAT"
    res.err(None, None, "unsupported_format",
            f"unsupported rail file format: {filename}")
    return res


__all__ = ["parse", "FORM11_ALIASES"]
