"""Shipping-line UPLOAD parsers — templates, byte readers, validation & mapping.

The reusable Data-Upload sub-module (module 4). Mirrors
:mod:`services.performance.upload_parsers`: pure functions that turn an uploaded
CSV/XLS/XLSX byte payload into a validated, mapped record set plus a preview and
user-friendly errors — WITHOUT touching the DB. The import step then hands the valid
records to the EXISTING :class:`services.shipping_lines.repository.ShippingLinesRepository.persist`
(same tables, same sha256 + row_sha256 idempotency).

Column mapping is ALIAS-DRIVEN (reuses :data:`services.shipping_lines.parsers.column_maps.ALIASES`),
so "Container No" / "Container Number" / "Cntr No" / "CNTR_NO" all map to one field.
Two document families:
  * IAL / EAL  -> advance-list containers  (map_container_row)  -> sl_advance_containers
  * EDO        -> flat delivery orders      (_map_edo_row)        -> sl_delivery_orders
"""
from __future__ import annotations

import csv
import datetime as _dt
import io
from typing import Any, Optional

from .parsers.column_maps import ALIASES, map_container_row, pick
from .parsers.common import (
    IST,
    clean,
    container_valid,
    norm_container,
    norm_header,
    norm_intish,
    to_num,
)

LIST_TYPES = ("IAL", "EAL", "EDO")
_TERMINALS = ("APMT", "BMCT", "GTI", "NSFT", "NSICT", "NSIGT")

# Normalised weight-header vocabulary for the required-column check (the actual
# value is resolved by common.resolve_weight during mapping).
_WEIGHT_HEADERS = (
    "grossweight", "grossweightinkgs", "grossweightinmt", "grosswgt", "grosswt",
    "vgm", "vgmweightinmt", "weight", "weightkg", "weightmt", "wt",
)

# EDO (delivery order) canonical field -> accepted normalised headers.
EDO_ALIASES: dict[str, tuple[str, ...]] = {
    "container_no": ALIASES["container_no"],
    "gate_pass_no": ("gatepassno", "gatepass", "gatepassnumber", "gpno", "passno", "gatepassno1"),
    "vehicle_no": ("vehicleno", "vehicle", "vehiclenumber", "truckno", "lorryno", "trailerno"),
    "shipping_agent_code": ("shippingagentcode", "shippingagent", "agentcode", "agent",
                            "cacode", "calinecode"),
    "iso_code": ALIASES["iso_code"],
    "equipment_status": ("equipmentstatus", "equipmentstatuscode", "status", "fclmty",
                         "fullempty"),
    "loading_port": ("loadingport", "pol", "portofloading", "loadport"),
    "dest_port": ("destport", "destinationport", "pod", "portofdischarge"),
    "final_pod": ("finalpod", "finalportofdischarge", "finaldestination"),
    "delivery_mode": ("deliverymode", "mode", "departuremode", "modeofdelivery"),
    "common_ref_number": ("commonrefnumber", "commonref", "referenceno", "refno"),
    "arrival_ts": ("arrivaldatetime", "arrival", "arrivaldate", "eta"),
    "receipt_date": ("receiptdate", "receipt", "gateindate"),
    "gate_pass_ts": ("gatepassdatetime", "gatepasstime"),
}

# ---------------------------------------------------------------- required contract
# canonical label shown to the user -> the alias tuple that satisfies it.
_REQUIRED = {
    "IAL": {
        "Container Number": ALIASES["container_no"],
        "ISO Code": ALIASES["iso_code"],
        "Gross Weight": _WEIGHT_HEADERS,
        "Shipping Line": ALIASES["shipping_line"],
        "Category": ALIASES["category_src"] + ALIASES["shipping_status"],
    },
    "EDO": {
        "Container Number": EDO_ALIASES["container_no"],
        "Gate Pass No": EDO_ALIASES["gate_pass_no"],
    },
}
_REQUIRED["EAL"] = _REQUIRED["IAL"]

# ---------------------------------------------------------------- templates
_ADVANCE_TEMPLATE_COLS = [
    "Container Number", "ISO Code", "Gross Weight", "Shipping Line", "Category",
    "Freight Kind", "POL", "POD", "Vessel Visit", "Bill of Lading", "Seal",
    "Reefer Temp", "IMDG", "UN Number", "Group", "Client", "Departure Mode",
    "Nominated CFS", "IEC Code", "GST No", "Commodity Code",
]
_EDO_TEMPLATE_COLS = [
    "Container Number", "Gate Pass No", "Vehicle No", "Shipping Agent", "ISO Code",
    "Equipment Status", "Loading Port", "Dest Port", "Final POD", "Delivery Mode",
    "Common Ref Number", "Arrival DateTime", "Receipt Date",
]
_SPECS = {
    "IAL": {"cols": _ADVANCE_TEMPLATE_COLS, "target": "advance",
            "example": ["BEAU2313280", "2210", "19880", "KMD", "IMPORT", "FULL", "LCH",
                        "INNSA", "KMIS0276", "KMC813961", "CF712911", "", "", "", "SBW",
                        "8GH", "T", "", "AAFCR2430M", "27AAFCR2430M1Z9", "GEN"]},
    "EAL": {"cols": _ADVANCE_TEMPLATE_COLS, "target": "advance",
            "example": ["SEGU9719798", "4532", "34010", "KMD", "EXPORT", "FULL", "INNSA",
                        "MYPKG", "KMIS0276", "", "CF921397", "-18", "", "", "DRT", "8GH",
                        "T", "", "", "", "GEN"]},
    "EDO": {"cols": _EDO_TEMPLATE_COLS, "target": "delivery",
            "example": ["SAJU2031655", "16494337", "MH43U7042", "UNF", "2210", "MTY",
                        "INNSA1", "THLCH", "THLCH", "G", "12062611183757",
                        "12062026:02:53", "12062026"]},
}


def list_type_ok(list_type: str) -> Optional[str]:
    lt = (list_type or "").strip().upper()
    return lt if lt in LIST_TYPES else None


def template_csv(list_type: str) -> str:
    spec = _SPECS[list_type]
    req = set(_REQUIRED[list_type].keys())
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(spec["cols"])
    # A '#'-prefixed guidance line — read_rows_from_bytes skips it on re-upload.
    w.writerow([f"# REQUIRED: {', '.join(req)} | all other columns are OPTIONAL. "
                f"Column names are flexible (e.g. 'Container No' / 'CNTR_NO' also work). "
                f"Delete this line and the example row before uploading."]
               + [""] * (len(spec["cols"]) - 1))
    w.writerow(spec["example"])
    return buf.getvalue()


def derive_terminal(filename: str) -> str:
    up = (filename or "").upper().replace(" ", "").replace("_", "")
    return next((t for t in _TERMINALS if t in up), "OTHER")


# ---------------------------------------------------------------- byte readers
def read_rows_from_bytes(content: bytes, filename: str) -> tuple[list[str], list[dict[str, Any]]]:
    """Return (header, rows) from a CSV / XLS / XLSX byte payload. Raises ValueError on
    an unreadable/empty file or unsupported extension."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            it = ws.iter_rows(values_only=True)
            try:
                header = [str(c).strip() if c is not None else "" for c in next(it)]
            except StopIteration:
                raise ValueError("empty_file")
            rows = []
            for values in it:
                if not any(v not in (None, "") for v in values):
                    continue
                rows.append({header[i]: (values[i] if i < len(values) else None)
                             for i in range(len(header))})
        finally:
            wb.close()
        return header, rows
    if name.endswith(".xls"):
        import xlrd
        book = xlrd.open_workbook(file_contents=content)
        sh = book.sheet_by_index(0)
        if sh.nrows == 0:
            raise ValueError("empty_file")
        header = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]
        rows = []
        for r in range(1, sh.nrows):
            values = [sh.cell_value(r, c) for c in range(sh.ncols)]
            if not any(str(v).strip() for v in values):
                continue
            rows.append({header[i]: values[i] for i in range(len(header))})
        return header, rows
    if name.endswith((".csv", ".txt")) or name == "":
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        all_rows = [r for r in reader if any((c or "").strip() for c in r)]
        if not all_rows:
            raise ValueError("empty_file")
        header = [c.strip() for c in all_rows[0]]
        rows = []
        for r in all_rows[1:]:
            if r and str(r[0]).strip().startswith("#"):   # skip template guidance lines
                continue
            rows.append({header[i]: (r[i] if i < len(r) else None) for i in range(len(header))})
        return header, rows
    raise ValueError("unsupported_format")


def is_codeco_xml_upload(header: list[str]) -> bool:
    """Detect the legacy EDO shape (CODECO XML embedded in a PAYLOAD cell) so the
    service can route it to the EXISTING parse_edo path instead of the flat mapper."""
    hset = {norm_header(h) for h in header}
    return "payload" in hset or "messagename" in hset


# ---------------------------------------------------------------- ParseResult
class ParseResult:
    def __init__(self) -> None:
        self.errors: list[dict[str, Any]] = []
        self.warnings: list[dict[str, Any]] = []
        self.records: list[dict[str, Any]] = []   # valid, mapped canonical rows
        self.preview: list[dict[str, Any]] = []
        self.row_count = 0
        self.invalid_count = 0
        self.duplicate_count = 0
        self.rejected = False                       # structural failure (wrong template)
        self.target = "advance"                     # 'advance' | 'delivery'

    def err(self, row: Optional[int], col: Optional[str], code: str, detail: str, raw: Any = None):
        self.errors.append({"row_number": row, "column_name": col, "error_code": code,
                            "error_detail": detail, "raw_value": (None if raw is None else str(raw))})

    def warn(self, row: Optional[int], col: Optional[str], code: str, detail: str):
        self.warnings.append({"row_number": row, "column_name": col, "error_code": code,
                              "error_detail": detail})


def check_required_columns(res: ParseResult, list_type: str, header: list[str]) -> bool:
    """Alias-aware required-column check. Missing → user-friendly error + rejected."""
    hset = {norm_header(h) for h in header if norm_header(h)}
    missing = []
    for label, aliases in _REQUIRED[list_type].items():
        if not any(a in hset for a in aliases):
            missing.append(label)
    if missing:
        for label in missing:
            res.err(None, label, "missing_column",
                    f"{label} column not found. Please download the latest template.")
        res.rejected = True
        return False
    return True


# ---------------------------------------------------------------- date coercion (EDO)
def _edo_dt(value: Any) -> Optional[_dt.datetime]:
    """Accept DDMMYYYY:HH:MM, a real datetime, or ISO — lenient → None."""
    if value is None:
        return None
    if isinstance(value, _dt.datetime):
        return value if value.tzinfo else value.replace(tzinfo=IST)
    from .parsers.common import parse_ddmmyyyy_time
    s = clean(value)
    if not s:
        return None
    d = parse_ddmmyyyy_time(s)
    if d is not None:
        return d
    try:
        return _dt.datetime.fromisoformat(s.replace("Z", "+00:00"))
    except ValueError:
        return None


def _edo_date(value: Any) -> Optional[_dt.date]:
    if isinstance(value, _dt.datetime):
        return value.date()
    if isinstance(value, _dt.date):
        return value
    from .parsers.common import parse_ddmmyyyy
    s = clean(value)
    if not s:
        return None
    d = parse_ddmmyyyy(s)
    if d is not None:
        return d
    try:
        return _dt.date.fromisoformat(s[:10])
    except ValueError:
        return None


def _map_edo_row(raw_row: dict[str, Any]) -> Optional[dict[str, Any]]:
    """Map one flat EDO row to the sl_delivery_orders canonical dict."""
    row_norm = {norm_header(k): v for k, v in raw_row.items() if norm_header(k)}

    def epick(field: str) -> Optional[str]:
        for a in EDO_ALIASES.get(field, ()):  # normalised
            if a in row_norm:
                v = clean(row_norm[a])
                if v is not None:
                    return v
        return None

    container_no = norm_container(epick("container_no"))
    if not container_no:
        return None
    return {
        "container_no": container_no,
        "container_valid_iso": container_valid(container_no),
        "gate_pass_no": epick("gate_pass_no"),
        "vehicle_no": epick("vehicle_no"),
        "shipping_agent_code": epick("shipping_agent_code"),
        "iso_code": norm_intish(epick("iso_code")),
        "equipment_status": epick("equipment_status"),
        "loading_port": epick("loading_port"),
        "dest_port": epick("dest_port"),
        "final_pod": epick("final_pod"),
        "delivery_mode": epick("delivery_mode"),
        "common_ref_number": epick("common_ref_number"),
        "arrival_ts": _edo_dt(row_norm.get("arrivaldatetime") or row_norm.get("arrival")),
        "receipt_date": _edo_date(row_norm.get("receiptdate") or row_norm.get("receipt")),
        "gate_pass_ts": _edo_dt(row_norm.get("gatepassdatetime")),
        # columns the flat template does not carry
        "document_number": None, "message_type": None, "sender_id": None,
        "receiving_party": None, "vcn": None, "imo_number": None, "call_sign": None,
        "stuff_destuff_flag": None, "vessel_country": None, "total_containers": None,
        "cargo_type": None, "gate_number": None, "ca_code": epick("shipping_agent_code"),
        "con_seal_status": None, "issued_ts": None, "raw_xml": None,
    }


# ---------------------------------------------------------------- main parse
def parse(list_type: str, header: list[str], rows: list[dict[str, Any]]) -> ParseResult:
    res = ParseResult()
    res.target = _SPECS[list_type]["target"]
    res.row_count = len(rows)
    if not check_required_columns(res, list_type, header):
        return res
    if res.target == "delivery":
        _parse_edo(res, rows)
    else:
        _parse_advance(res, list_type, rows)
    return res


def _parse_advance(res: ParseResult, list_type: str, rows: list[dict[str, Any]]) -> None:
    terminal = "OTHER"
    seen: set[tuple[str, str]] = set()
    for i, raw in enumerate(rows, start=1):
        row_norm = {norm_header(k): v for k, v in raw.items() if norm_header(k)}
        cn = norm_container(pick(row_norm, "container_no"))
        if not cn:
            res.err(i, "Container Number", "empty_container", "container number is empty", None)
            res.invalid_count += 1
            continue
        # required cells
        iso = pick(row_norm, "iso_code")
        line = pick(row_norm, "shipping_line")
        cat = pick(row_norm, "category_src") or pick(row_norm, "shipping_status")
        from .parsers.common import resolve_weight
        wt_kg, _uom = resolve_weight(row_norm)
        missing = []
        if not iso:
            missing.append("ISO Code")
        if not line:
            missing.append("Shipping Line")
        if not cat:
            missing.append("Category")
        # weight: distinguish "empty" (missing) from "present but unparseable"
        raw_wt = next((row_norm[h] for h in _WEIGHT_HEADERS if h in row_norm and clean(row_norm[h])), None)
        if raw_wt is None:
            missing.append("Gross Weight")
        elif wt_kg is None:
            res.err(i, "Gross Weight", "invalid_weight",
                    f"weight '{clean(raw_wt)}' is not a valid number", raw_wt)
            res.invalid_count += 1
            continue
        if missing:
            res.err(i, ", ".join(missing), "empty_required",
                    f"required value(s) empty: {', '.join(missing)}")
            res.invalid_count += 1
            continue
        # soft checks (importable, flagged as warnings)
        if not container_valid(cn):
            res.warn(i, "Container Number", "container_iso6346_invalid",
                     f"{cn} fails the ISO-6346 check digit (imported, flagged)")
        # duplicate within this file
        key = (cn, pick(row_norm, "vessel_visit") or "")
        if key in seen:
            res.duplicate_count += 1
            res.warn(i, "Container Number", "duplicate_in_file",
                     f"{cn} already appears earlier in this file (skipped)")
            continue
        seen.add(key)
        mapped = map_container_row(raw, list_type=list_type, terminal=terminal)
        if mapped is not None:
            res.records.append(mapped)
    res.preview = [_advance_preview(r) for r in res.records[:20]]


def _advance_preview(r: dict[str, Any]) -> dict[str, Any]:
    return {"Container": r["container_no"], "List": r["list_type"], "ISO": r.get("iso_code"),
            "Line": r.get("shipping_line_code"), "Category": r.get("category"),
            "Freight": r.get("freight_kind"),
            "Weight(kg)": r.get("gross_weight_kg"), "POD": r.get("pod"),
            "BL": r.get("bill_of_lading")}


def _parse_edo(res: ParseResult, rows: list[dict[str, Any]]) -> None:
    seen: set[tuple[str, str]] = set()
    for i, raw in enumerate(rows, start=1):
        row_norm = {norm_header(k): v for k, v in raw.items() if norm_header(k)}
        mapped = _map_edo_row(raw)
        if mapped is None:
            res.err(i, "Container Number", "empty_container", "container number is empty")
            res.invalid_count += 1
            continue
        if not mapped.get("gate_pass_no"):
            res.err(i, "Gate Pass No", "empty_required", "gate pass number is empty")
            res.invalid_count += 1
            continue
        if not container_valid(mapped["container_no"]):
            res.warn(i, "Container Number", "container_iso6346_invalid",
                     f"{mapped['container_no']} fails the ISO-6346 check digit (imported, flagged)")
        key = (mapped["container_no"], mapped.get("gate_pass_no") or "")
        if key in seen:
            res.duplicate_count += 1
            res.warn(i, "Container Number", "duplicate_in_file",
                     f"{mapped['container_no']} / gate pass {mapped.get('gate_pass_no')} duplicated (skipped)")
            continue
        seen.add(key)
        res.records.append(mapped)
    res.preview = [{"Container": r["container_no"], "GatePass": r.get("gate_pass_no"),
                    "Vehicle": r.get("vehicle_no"), "Agent": r.get("shipping_agent_code"),
                    "ISO": r.get("iso_code"), "Status": r.get("equipment_status")}
                   for r in res.records[:20]]
