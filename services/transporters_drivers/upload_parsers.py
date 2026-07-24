"""Transporters & Drivers UPLOAD parsers — templates, byte readers, validation & mapping.

The reusable Data-Upload sub-module for the Transporter & Driver masters. Mirrors
:mod:`services.cfs_ecy.upload_parsers`: pure functions that turn an uploaded
CSV/XLS/XLSX byte payload into a validated, mapped record set plus a preview and
user-friendly errors — WITHOUT touching the DB. The import step then hands the valid
records to :class:`services.transporters_drivers.repository.TransportersDriversRepository`,
which upserts them into the EXISTING masters:
  * TRANSPORTER -> core.transporter      (upsert on source_company_id, the same key
    scripts/import_transporter_master.py uses -> idempotent, duplicate-safe).
  * DRIVER      -> core.driver      (upsert on licence_no_norm, the same key
    scripts/import_driver_master.py uses).

Column mapping is ALIAS-DRIVEN (header is normalised, then matched against an alias
table), so "Transporter Name" / "Company Name" / "Transporter_Name" all map to one
field and "Driver Name" / "Driver_Name" / "Name" all map to another. The entity type
(TRANSPORTER / DRIVER) is supplied by the upload's selector — it is not a column.
"""
from __future__ import annotations

import csv
import datetime as _dt
import io
import re
from typing import Any, Optional

ENTITIES = ("TRANSPORTER", "DRIVER")

# JNPA operates in IST; date-only master fields (licence_valid_to, dob) carry no time.
_VALID_STATUS = {
    "TRANSPORTER": {"ACTIVE", "SUSPENDED", "BLACKLISTED"},
    "DRIVER": {"ACTIVE", "INACTIVE"},
}

# ---------------------------------------------------------------- column aliases
# canonical field -> accepted NORMALISED header names (see norm_header). First
# present, non-empty match wins.
TRANSPORTER_ALIASES: dict[str, tuple[str, ...]] = {
    "source_company_id": (
        "companyid", "company", "transporterid", "transportercode1", "sourcecompanyid",
        "compid", "cmpid", "id",
    ),
    "name": (
        "companyname", "transportername", "name", "transporter", "firmname",
        "organisationname", "organizationname", "tptname",
    ),
    "code": (
        "transportercode", "code", "licensecode", "licencecode", "regcode", "tptcode",
    ),
    "gstin": ("gstin", "gst", "gstno", "gstnumber", "gstinnumber"),
    "contact_person": (
        "contactperson", "contactpersonname", "contactname", "personname", "poc",
        "primarycontact",
    ),
    "designation": ("designation", "role", "title", "post"),
    "email": ("email", "emailid", "emailaddress", "mail", "contactemail"),
    "mobile": (
        "mobile", "mobileno", "mobilenumber", "phone", "phoneno", "phonenumber",
        "contactnumber", "contactno", "contact",
    ),
    "address": ("address", "addr", "officeaddress", "location", "fulladdress"),
    "status": ("status", "state", "active"),
    "source_user_id": ("useruserid", "userid", "sourceuserid"),
}

DRIVER_ALIASES: dict[str, tuple[str, ...]] = {
    "licence_no": (
        "licencenumber", "licenceno", "licensenumber", "licenseno", "dlnumber", "dlno",
        "drivinglicence", "drivinglicense", "licence", "license",
    ),
    "name": (
        "drivername", "name", "fullname", "driver", "driverfullname",
    ),
    "company_name": (
        "companyname", "transportername", "company", "transporter", "firmname",
        "employer",
    ),
    "licence_type": (
        "licencetype", "licensetype", "dltype", "type", "vehicleclass", "class",
    ),
    "licence_valid_to": (
        "licencevalidto", "licenseexpiry", "licencevalidity", "validity", "validupto",
        "validtill", "expiry", "expirydate", "validto",
    ),
    "dob": ("dob", "dateofbirth", "birthdate", "birthday"),
    "latest_pdp_number": (
        "latestpdpnumber", "pdpnumber", "pdpno", "pdp", "permitnumber",
    ),
    "status": ("status", "state", "active"),
}

# canonical label shown to the user -> the alias tuple that satisfies it (a required
# column). Entity-specific.
_REQUIRED = {
    "TRANSPORTER": {
        "Company ID": TRANSPORTER_ALIASES["source_company_id"],
        "Company Name": TRANSPORTER_ALIASES["name"],
    },
    "DRIVER": {
        "Licence Number": DRIVER_ALIASES["licence_no"],
        "Driver Name": DRIVER_ALIASES["name"],
    },
}

# ---------------------------------------------------------------- templates
_TEMPLATES: dict[str, dict[str, Any]] = {
    "TRANSPORTER": {
        "columns": ["Company ID", "Company Name", "Transporter Code", "GSTIN",
                    "Contact Person", "Designation", "Email", "Mobile", "Address",
                    "Status"],
        "example": ["100245", "Bharat Transport Pvt Ltd", "TPT-100245",
                    "27AABCB1234C1ZV", "Ramesh Kumar", "Fleet Manager",
                    "ops@bharattransport.in", "9876543210",
                    "Plot 14, JNPT Road, Navi Mumbai", "ACTIVE"],
        "guidance": (
            "# REQUIRED: Company ID (unique integer key), Company Name. OPTIONAL: "
            "Transporter Code, GSTIN, Contact Person, Designation, Email, Mobile, "
            "Address, Status (ACTIVE/SUSPENDED/BLACKLISTED, default ACTIVE). "
            "Company ID is the idempotency key — re-uploading updates that transporter. "
            "Mobile = 10 digits. Column names are flexible (e.g. 'Transporter Name', "
            "'Transporter_Name' also map). Delete this line and the example row "
            "before uploading."
        ),
    },
    "DRIVER": {
        "columns": ["Licence Number", "Driver Name", "Company Name", "Licence Type",
                    "Licence Valid To", "DOB", "Latest PDP Number", "Status"],
        "example": ["MH0120220001234", "Suresh Patil", "Bharat Transport Pvt Ltd",
                    "HMV", "31/12/2027", "15/06/1985", "PDP-2024-04512", "ACTIVE"],
        "guidance": (
            "# REQUIRED: Licence Number (unique key), Driver Name. OPTIONAL: Company "
            "Name (maps the driver to a Transporter by name), Licence Type "
            "(default HMV), Licence Valid To (DD/MM/YYYY), DOB (DD/MM/YYYY), Latest "
            "PDP Number, Status (ACTIVE/INACTIVE, default ACTIVE). Licence Number is "
            "the idempotency key — re-uploading updates that driver. Column names are "
            "flexible (e.g. 'Driver Name', 'Name', 'DL Number' also map). Delete this "
            "line and the example row before uploading."
        ),
    },
}

_GST_RE = re.compile(r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][0-9A-Z]Z[0-9A-Z]$")
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_DATE_FORMATS = ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d %b %Y", "%d-%b-%Y")


def entity_ok(value: Optional[str]) -> Optional[str]:
    v = (value or "").strip().upper()
    return v if v in ENTITIES else None


def norm_header(name: Any) -> str:
    """Lowercase and drop everything that is not a letter or digit so header
    variations ('Company Name', 'Transporter_Name', 'CompanyName') collapse to one key."""
    return re.sub(r"[^a-z0-9]", "", str(name).strip().lower())


def clean(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    v = str(raw).strip()
    return v or None


def norm_licence(raw: Any) -> Optional[str]:
    """UPPER + alphanumeric-only, matching scripts/import_driver_master.norm_licence."""
    v = clean(raw)
    return re.sub(r"[^A-Z0-9]", "", v.upper()) if v else None


def clean_mobile(raw: Any) -> Optional[str]:
    """Return a 10-digit mobile, stripping 91/0 country/trunk prefixes (mirrors
    scripts/import_transporter_master.clean_mobile). None if it cannot be normalised."""
    v = clean(raw)
    if not v:
        return None
    digits = re.sub(r"\D", "", v)
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    elif len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]
    return digits if len(digits) == 10 else None


def parse_date(raw: Any) -> Optional[_dt.date]:
    """Parse a DD/MM/YYYY-family or ISO date. Accepts an Excel native date/datetime.
    None if unrecognised."""
    if raw is None:
        return None
    if isinstance(raw, _dt.datetime):
        return raw.date()
    if isinstance(raw, _dt.date):
        return raw
    s = clean(raw)
    if not s:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return _dt.date.fromisoformat(s[:10])
    except ValueError:
        return None


def norm_status(raw: Any, entity: str, default: str = "ACTIVE") -> tuple[Optional[str], bool]:
    """Return (status, ok). Blank -> default. Unknown -> (raw_upper, False)."""
    v = (clean(raw) or "").upper()
    if not v:
        return default, True
    if v in _VALID_STATUS.get(entity, set()):
        return v, True
    return v, False


def template_csv(entity: str) -> str:
    spec = _TEMPLATES[entity]
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(spec["columns"])
    guidance_row = [spec["guidance"]] + [""] * (len(spec["columns"]) - 1)
    w.writerow(guidance_row)
    w.writerow(spec["example"])
    return buf.getvalue()


# ---------------------------------------------------------------- byte readers
def read_rows_from_bytes(content: bytes, filename: str) -> tuple[list[str], list[dict[str, Any]]]:
    """Return (header, rows) from a CSV / XLS / XLSX byte payload. Raises ValueError on
    an unreadable/empty file or unsupported extension. Mirrors the CFS-ECY reader."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            it = ws.iter_rows(values_only=True)
            try:
                header = [str(c).strip() if c is not None else "" for c in next(it)]
            except StopIteration:
                raise ValueError("empty_file")
            rows = []
            for values in it:
                if not any(v not in (None, "") for v in values):
                    continue
                rows.append({header[i]: (values[i] if i < len(values) else None)
                             for i in range(len(header))})
        finally:
            wb.close()
        return header, rows
    if name.endswith(".xls"):
        import xlrd
        book = xlrd.open_workbook(file_contents=content)
        sh = book.sheet_by_index(0)
        if sh.nrows == 0:
            raise ValueError("empty_file")
        header = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]
        rows = []
        for r in range(1, sh.nrows):
            values = [sh.cell_value(r, c) for c in range(sh.ncols)]
            if not any(str(v).strip() for v in values):
                continue
            rows.append({header[i]: values[i] for i in range(len(header))})
        return header, rows
    if name.endswith((".csv", ".txt")) or name == "":
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        all_rows = [r for r in reader if any((c or "").strip() for c in r)]
        if not all_rows:
            raise ValueError("empty_file")
        header = [c.strip() for c in all_rows[0]]
        rows = []
        for r in all_rows[1:]:
            if r and str(r[0]).strip().startswith("#"):   # skip template guidance lines
                continue
            rows.append({header[i]: (r[i] if i < len(r) else None) for i in range(len(header))})
        return header, rows
    raise ValueError("unsupported_format")


# ---------------------------------------------------------------- ParseResult
class ParseResult:
    def __init__(self) -> None:
        self.errors: list[dict[str, Any]] = []
        self.warnings: list[dict[str, Any]] = []
        self.records: list[dict[str, Any]] = []   # valid, mapped canonical rows
        self.preview: list[dict[str, Any]] = []
        self.row_count = 0
        self.invalid_count = 0
        self.duplicate_count = 0
        self.rejected = False                       # structural failure (wrong template)

    def err(self, row: Optional[int], col: Optional[str], code: str, detail: str, raw: Any = None):
        self.errors.append({"row_number": row, "column_name": col, "error_code": code,
                            "error_detail": detail, "raw_value": (None if raw is None else str(raw))})

    def warn(self, row: Optional[int], col: Optional[str], code: str, detail: str):
        self.warnings.append({"row_number": row, "column_name": col, "error_code": code,
                              "error_detail": detail})


def check_required_columns(res: ParseResult, header: list[str], entity: str) -> bool:
    """Alias-aware required-column check. Missing -> user-friendly error + rejected."""
    hset = {norm_header(h) for h in header if norm_header(h)}
    missing = []
    for label, aliases in _REQUIRED[entity].items():
        if not any(a in hset for a in aliases):
            missing.append(label)
    if missing:
        for label in missing:
            res.err(None, label, "missing_column",
                    f"{label} column not found. Please download the latest template.")
        res.rejected = True
        return False
    return True


def _pick(row_norm: dict[str, Any], aliases: dict[str, tuple[str, ...]], canonical: str) -> Optional[str]:
    for src in aliases.get(canonical, ()):
        if src in row_norm:
            v = clean(row_norm[src])
            if v is not None:
                return v
    return None


# ---------------------------------------------------------------- main parse
def parse(header: list[str], rows: list[dict[str, Any]], *, entity: str,
          source_file: Optional[str] = None) -> ParseResult:
    """Validate + map uploaded rows for one file of the given entity type."""
    entity = (entity or "").upper()
    if entity not in ENTITIES:
        res = ParseResult(); res.rejected = True
        res.err(None, None, "invalid_entity", f"unknown upload type '{entity}'")
        return res
    res = ParseResult()
    res.row_count = len(rows)
    if not check_required_columns(res, header, entity):
        return res
    if entity == "TRANSPORTER":
        return _parse_transporters(res, rows, source_file)
    return _parse_drivers(res, rows, source_file)


def _parse_transporters(res: ParseResult, rows: list[dict[str, Any]],
                        source_file: Optional[str]) -> ParseResult:
    A = TRANSPORTER_ALIASES
    seen: set[int] = set()
    for i, raw in enumerate(rows, start=1):
        row_norm = {norm_header(k): v for k, v in raw.items() if norm_header(k)}

        raw_cid = _pick(row_norm, A, "source_company_id")
        cid_digits = re.sub(r"\D", "", raw_cid) if raw_cid else ""
        if not cid_digits:
            res.err(i, "Company ID", "invalid_company_id",
                    f"Company ID '{raw_cid}' is empty or not an integer", raw_cid)
            res.invalid_count += 1
            continue
        company_id = int(cid_digits)

        name = _pick(row_norm, A, "name")
        if not name:
            res.err(i, "Company Name", "empty_required", "Company Name is empty")
            res.invalid_count += 1
            continue

        if company_id in seen:
            res.duplicate_count += 1
            res.warn(i, "Company ID", "duplicate_in_file",
                     f"Company ID {company_id} already appears earlier in this file (skipped)")
            continue
        seen.add(company_id)

        # optional fields + soft format checks (warnings — the row still imports)
        gstin = _pick(row_norm, A, "gstin")
        if gstin and not _GST_RE.match(gstin.upper()):
            res.warn(i, "GSTIN", "gstin_format_invalid",
                     f"GSTIN '{gstin}' does not match the 15-char GST format (imported, flagged)")
        email = _pick(row_norm, A, "email")
        if email and not _EMAIL_RE.match(email):
            res.warn(i, "Email", "email_format_invalid",
                     f"email '{email}' looks invalid (imported without email)")
            email = None
        raw_mobile = _pick(row_norm, A, "mobile")
        mobile = clean_mobile(raw_mobile)
        if raw_mobile and mobile is None:
            res.warn(i, "Mobile", "mobile_invalid",
                     f"mobile '{raw_mobile}' is not a 10-digit number (imported without mobile)")
        status, ok = norm_status(_pick(row_norm, A, "status"), "TRANSPORTER")
        if not ok:
            res.warn(i, "Status", "status_invalid",
                     f"status '{status}' is not ACTIVE/SUSPENDED/BLACKLISTED (defaulted to ACTIVE)")
            status = "ACTIVE"

        raw_uid = _pick(row_norm, A, "source_user_id")
        uid_digits = re.sub(r"\D", "", raw_uid) if raw_uid else ""

        res.records.append({
            "source_company_id": company_id,
            "source_user_id": int(uid_digits) if uid_digits else None,
            "name": name,
            "code": _pick(row_norm, A, "code"),
            "gstin": (gstin.upper() if gstin else None),
            "contact_person": _pick(row_norm, A, "contact_person"),
            "designation": _pick(row_norm, A, "designation"),
            "email": email,
            "mobile": mobile,
            "address": _pick(row_norm, A, "address"),
            "status": status,
        })

    res.preview = [{
        "Company ID": r["source_company_id"],
        "Company Name": r["name"],
        "Code": r["code"] or "—",
        "GSTIN": r["gstin"] or "—",
        "Mobile": r["mobile"] or "—",
        "Status": r["status"],
    } for r in res.records[:20]]
    return res


def _parse_drivers(res: ParseResult, rows: list[dict[str, Any]],
                   source_file: Optional[str]) -> ParseResult:
    A = DRIVER_ALIASES
    seen: set[str] = set()
    for i, raw in enumerate(rows, start=1):
        row_norm = {norm_header(k): v for k, v in raw.items() if norm_header(k)}

        raw_lic = _pick(row_norm, A, "licence_no")
        lic_norm = norm_licence(raw_lic)
        if not lic_norm:
            res.err(i, "Licence Number", "empty_licence", "Licence Number is empty")
            res.invalid_count += 1
            continue
        if len(lic_norm) < 5:
            res.err(i, "Licence Number", "licence_too_short",
                    f"Licence Number '{raw_lic}' is too short to be valid", raw_lic)
            res.invalid_count += 1
            continue

        name = _pick(row_norm, A, "name")
        if not name:
            res.err(i, "Driver Name", "empty_required", "Driver Name is empty")
            res.invalid_count += 1
            continue

        if lic_norm in seen:
            res.duplicate_count += 1
            res.warn(i, "Licence Number", "duplicate_in_file",
                     f"Licence {lic_norm} already appears earlier in this file (skipped)")
            continue
        seen.add(lic_norm)

        company_name = _pick(row_norm, A, "company_name")
        if not company_name:
            res.warn(i, "Company Name", "no_transporter_mapping",
                     "no Company Name — driver imported without a transporter mapping")

        raw_valid = _pick(row_norm, A, "licence_valid_to")
        licence_valid_to = parse_date(raw_valid)
        if raw_valid and licence_valid_to is None:
            res.warn(i, "Licence Valid To", "invalid_date",
                     f"Licence Valid To '{raw_valid}' is not a recognised date (imported blank)")
        raw_dob = _pick(row_norm, A, "dob")
        dob = parse_date(raw_dob)
        if raw_dob and dob is None:
            res.warn(i, "DOB", "invalid_date",
                     f"DOB '{raw_dob}' is not a recognised date (imported blank)")
        status, ok = norm_status(_pick(row_norm, A, "status"), "DRIVER")
        if not ok:
            res.warn(i, "Status", "status_invalid",
                     f"status '{status}' is not ACTIVE/INACTIVE (defaulted to ACTIVE)")
            status = "ACTIVE"

        res.records.append({
            "licence_no": clean(raw_lic),
            "licence_no_norm": lic_norm,
            "name": name,
            "company_name": company_name,
            "licence_type": _pick(row_norm, A, "licence_type") or "HMV",
            "licence_valid_to": licence_valid_to,
            "dob": dob,
            "latest_pdp_number": _pick(row_norm, A, "latest_pdp_number"),
            "status": status,
        })

    res.preview = [{
        "Licence": r["licence_no_norm"],
        "Driver Name": r["name"],
        "Company": r["company_name"] or "—",
        "Type": r["licence_type"],
        "Valid To": (r["licence_valid_to"].strftime("%d/%m/%Y") if r["licence_valid_to"] else "—"),
        "Status": r["status"],
    } for r in res.records[:20]]
    return res
